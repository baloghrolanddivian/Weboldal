from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import subprocess
import zipfile
from collections import defaultdict, deque
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Sequence

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except Exception:  # pragma: no cover - optional dependency handling
    openpyxl = None
    PatternFill = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency handling
    PdfReader = None


FIELDNAMES = [
    "termek",
    "szin",
    "meret",
    "m2",
    "db",
    "ossz_m2",
    "egyseg_ar",
    "netto_ar",
    "kod",
]

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
RUNTIME_DIR = BASE_DIR / "runtime" / "nettfront"
DEFAULT_TRANSLATIONS = DATA_DIR / "nettfront-translations.json"
DEFAULT_ALKATRESZEK = DATA_DIR / "nettfront-alkatreszek.xlsx"

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if PatternFill else None
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if PatternFill else None

TEXT_FIXES = {
    "MagasfĂ©nyĂ» antracit": "Magasfényű antracit",
    "MagasfĂ©nyĂ» krĂ©m": "Magasfényű krém",
    "MagasfĂ©nyĂ» latte": "Magasfényű latte",
    "MagasfĂ©nyĂ» fehĂ©r": "Magasfényű fehér",
    "Matt ANTRAZITE": "Matt antracit",
}

MOJIBAKE_FIXES = {
    "â”śĂ­": "á",
    "â”śÄ™": "é",
    "â”śĹź": "í",
    "â”śâ”‚": "ó",
    "â”śĂ‚": "ö",
    "â”ĽÄą": "ő",
    "â”śâ•‘": "ú",
    "â”śâ•ť": "ü",
    "â”Ľâ–“": "ű",
    "â”śĂĽ": "Á",
    "â”śĂ«": "É",
    "â”śĹą": "Í",
    "â”śĂ´": "Ó",
    "â”śÄľ": "Ö",
    "â”ĽĂ‰": "Ő",
    "â”śĂś": "Ú",
    "â”śĹĄ": "Ü",
    "â”Ľâ–‘": "Ű",
}


@dataclass
class ProcurementArtifacts:
    invoice_rows: list[dict[str, str]]
    invoice_csv: bytes
    procurement_csv: bytes
    missing_codes: list[str]


@dataclass
class CompareArtifacts:
    invoice_rows: list[dict[str, str]]
    invoice_csv: bytes
    compare_workbook: bytes
    order_row_count: int


@dataclass
class NettfrontArtifacts:
    invoice_rows: list[dict[str, str]]
    invoice_csv: bytes
    procurement_csv: bytes
    compare_workbook: bytes | None
    missing_codes: list[str]
    order_uploaded: bool
    order_row_count: int


def _require_pdf_support() -> None:
    if PdfReader is None:
        raise RuntimeError("A pypdf csomag nincs telepítve.")


def _require_workbook_support() -> None:
    if openpyxl is None:
        raise RuntimeError("Az openpyxl csomag nincs telepítve.")


def extract_page_text(pages: Iterable) -> str:
    return "\n".join((page.extract_text() or "").strip() for page in pages)


def normalize_numeric(value: str) -> str:
    return value.replace("\u00a0", " ").replace(" ", "").replace(",", ".")


def is_numeric(value: str) -> bool:
    try:
        float(normalize_numeric(value))
        return True
    except ValueError:
        return False


def format_currency(value: str) -> str:
    return normalize_numeric(value).replace(".", ",")


def fix_mojibake(value: str) -> str:
    fixed = value
    for bad, good in MOJIBAKE_FIXES.items():
        fixed = fixed.replace(bad, good)
    return fixed


def normalize_text(value: str) -> str:
    fixed = fix_mojibake(value)
    fixed = fixed.replace("\u00a0", " ").replace("  ", " ").strip()
    fixed = fixed.replace("Ă»", "ű").replace("Ă‚", "")
    fixed = " ".join(fixed.split())
    return TEXT_FIXES.get(fixed, fixed)


def detect_handedness(*values: str) -> str | None:
    normalized_values = [normalize_text(value).upper() for value in values]
    joined = " ".join(normalized_values)
    token_map = {
        " JOBBOS": "J",
        " BALOS": "B",
        " JOBB": "J",
        " BAL": "B",
    }

    for token, code in token_map.items():
        if token in f" {joined} ":
            return code

    compact_values = [value.replace(" ", "") for value in normalized_values]
    for value in compact_values:
        if value.endswith(("_J", "-J", "J")):
            return "J"
        if value.endswith(("_B", "-B", "B")):
            return "B"
    return None


def strip_handedness(value: str) -> str:
    normalized = normalize_text(value)
    for token in (" jobbos", " balos", " jobb", " bal"):
        if normalized.lower().endswith(token):
            normalized = normalized[: -len(token)]
            break
    return normalize_text(normalized.rstrip("_- /"))


def parse_rows(text: str) -> list[dict[str, str]]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    rows: list[dict[str, str]] = []
    in_table = False
    i = 0

    while i < len(lines):
        line = lines[i]

        if line == "NettĂł Ăˇr" or line == "Nettó ár":
            in_table = True
            i += 1
            continue

        if line.startswith("Ă–sszesĂ­tve") or line.startswith("Összesítve"):
            in_table = False

        if in_table and line.isdigit() and i + 8 < len(lines):
            termek, szin, meret = lines[i + 1 : i + 4]
            m2 = lines[i + 4]
            db = lines[i + 5]
            ossz_m2 = lines[i + 6]
            egyseg_ar = lines[i + 7]
            netto_ar = lines[i + 8]

            numeric_values = (m2, db, ossz_m2, egyseg_ar, netto_ar)
            if "x" in meret and all(is_numeric(value) for value in numeric_values):
                rows.append(
                    {
                        "termek": termek,
                        "szin": szin,
                        "meret": meret,
                        "m2": normalize_numeric(m2),
                        "db": normalize_numeric(db),
                        "ossz_m2": normalize_numeric(ossz_m2),
                        "egyseg_ar": normalize_numeric(egyseg_ar),
                        "netto_ar": normalize_numeric(netto_ar),
                    }
                )
                i += 9
                if i < len(lines) and lines[i].startswith("NĂ©gyzetmĂ©terĂˇr"):
                    i += 1
                continue

        i += 1

    return rows


def load_translations(path: Path = DEFAULT_TRANSLATIONS) -> dict:
    if not path.is_file():
        raise FileNotFoundError(f"Translation table not found: {path}")
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _normalize_product_code_for_size(product_code: str, meret: str) -> str:
    if not product_code or "_" not in product_code:
        return product_code

    prefix, suffix = product_code.split("_", 1)
    if prefix == "NFAH" and meret != "718x250":
        return f"NFA_{suffix}"
    if prefix == "NFAL" and meret != "357x100":
        return f"NFA_{suffix}"
    if prefix == "NFAIA" and not meret.startswith("655x397"):
        return f"NFA_{suffix}"
    return product_code


def apply_translations(rows: list[dict[str, str]], table: dict) -> list[dict[str, str]]:
    products = table.get("products", {})
    colors = table.get("colors", {})
    standard_sizes = set(table.get("standard_sizes", []))
    translated: list[dict[str, str]] = []
    arched_fallback_index: dict[tuple[str, str, str], int] = {}

    for row in rows:
        raw_termek = normalize_text(row["termek"])
        termek = strip_handedness(raw_termek)
        szin = normalize_text(row["szin"])
        raw_meret = normalize_text(row["meret"])
        meret = raw_meret.replace(" ", "")

        product_meta = products.get(termek, {})
        color_meta = colors.get(szin, {})
        product_code = _normalize_product_code_for_size(product_meta.get("code", termek), meret)
        color_code = color_meta.get("code", szin)
        model_code = product_code.split("_")[-1]
        is_standard_size = meret in standard_sizes

        if model_code == "LU" and color_meta.get("name") == "Dune Beige":
            color_code = "KAFS"
        if model_code == "LU" and color_meta.get("name") == "Cedar Green":
            color_code = "PRAS"

        handedness = detect_handedness(raw_termek, raw_meret)
        is_arched_size = meret.startswith("655x397")

        if is_arched_size:
            if not handedness:
                pair_key = (model_code, color_code, "655x397")
                current_index = arched_fallback_index.get(pair_key, 0)
                handedness = "J" if current_index % 2 == 0 else "B"
                arched_fallback_index[pair_key] = current_index + 1
            kod = "_".join(filter(None, ["NFAIA", model_code, color_code, "655x397", handedness]))
            meret = "655x397"
        elif meret == "718x250":
            kod = "_".join(filter(None, ["NFAH", model_code, color_code, meret]))
        elif meret == "357x100":
            kod = "_".join(filter(None, ["NFAL", model_code, color_code, meret]))
        elif is_standard_size:
            kod = "_".join(filter(None, [product_code, color_code, meret]))
        else:
            kod = "_".join(filter(None, ["NFAY", model_code, color_code, meret]))

        translated.append(
            {
                **row,
                "termek": product_meta.get("name", termek),
                "szin": color_meta.get("name", szin),
                "meret": meret,
                "kod": kod,
            }
        )

    return translated


def left_until_underscore(value: object) -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        return ""

    parts = text.split("_")
    if len(parts) >= 5 and parts[0] == "NFAIA" and parts[3] == "655x397" and parts[4] in {"J", "B"}:
        return "_".join(parts[:5])

    idx = text.find("_", 15)
    if idx == -1:
        return text
    return text[:idx]


def load_alkatresz_map(path: Path = DEFAULT_ALKATRESZEK) -> dict[str, str]:
    _require_workbook_support()
    if not path.is_file():
        raise FileNotFoundError(f"Alkatreszek file not found: {path}")

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    mapping: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[0] if row else None
        if not value:
            continue
        key = left_until_underscore(value)
        mapping.setdefault(key, value)
    return mapping


def _extract_alkatresz_code(value: object) -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        return ""

    match = re.search(r"\bNFA[A-Z0-9_]+\b", text, re.IGNORECASE)
    if not match:
        return ""
    return match.group(0)


def _load_alkatresz_rows_from_csv_bytes(data: bytes) -> list[list[str]]:
    encoding = "utf-8-sig"
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        encoding = "utf-16"

    text = data.decode(encoding, errors="ignore")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.get_dialect("excel")

    reader = csv.reader(io.StringIO(text), dialect)
    return [[str(cell).strip() for cell in row] for row in reader]


def load_alkatresz_map_from_bytes(data: bytes, file_name: str = "") -> dict[str, str]:
    rows: list[Sequence[object]]
    suffix = Path(file_name).suffix.lower()

    if suffix in {".xlsx", ".xlsm"} or zipfile.is_zipfile(io.BytesIO(data)):
        _require_workbook_support()
        workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    else:
        rows = _load_alkatresz_rows_from_csv_bytes(data)

    mapping: dict[str, str] = {}
    for row in rows:
        if not row:
            continue
        for value in row:
            code = _extract_alkatresz_code(value)
            if not code:
                continue
            mapping[left_until_underscore(code)] = code
            break

    if not mapping:
        raise ValueError("A feltöltött alkatrészlistában nem találtam beolvasható kódokat.")
    return mapping


def build_invoice_rows(pdf_bytes: bytes) -> list[dict[str, str]]:
    _require_pdf_support()
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = extract_page_text(reader.pages)
    rows = parse_rows(text)
    rows = apply_translations(rows, load_translations())

    for row in rows:
        row["egyseg_ar"] = format_currency(row["egyseg_ar"])
        row["netto_ar"] = format_currency(row["netto_ar"])

    return rows


def build_invoice_csv(rows: list[dict[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=FIELDNAMES, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


COLOR_FALLBACK_CODES = {
    "PRA": "PRAS",
    "PRAS": "PRA",
    "KAF": "KAFS",
    "KAFS": "KAF",
    "BGA": "BGAS",
    "BGAS": "BGA",
    "GFA": "GFAS",
    "GFAS": "GFA",
    "FEA": "FEAS",
    "FEAS": "FEA",
}


def _procurement_code_fallbacks(kod: str) -> list[str]:
    candidates: list[str] = []

    if kod.startswith("NFAY_"):
        candidates.append("NFA_" + kod[len("NFAY_") :])
    elif kod.startswith("NFA_"):
        candidates.append("NFAY_" + kod[len("NFA_") :])

    parts = kod.split("_")
    if len(parts) >= 4:
        color_code = parts[2]
        fallback_color = COLOR_FALLBACK_CODES.get(color_code)
        if fallback_color:
            swapped_parts = parts[:]
            swapped_parts[2] = fallback_color
            swapped_kod = "_".join(swapped_parts)
            candidates.append(swapped_kod)

            if swapped_kod.startswith("NFAY_"):
                candidates.append("NFA_" + swapped_kod[len("NFAY_") :])
            elif swapped_kod.startswith("NFA_"):
                candidates.append("NFAY_" + swapped_kod[len("NFA_") :])

    unique_candidates: list[str] = []
    for candidate in candidates:
        if candidate and candidate != kod and candidate not in unique_candidates:
            unique_candidates.append(candidate)
    return unique_candidates


def build_procurement_csv(rows: list[dict[str, str]], alkatresz_map: dict[str, str] | None = None) -> tuple[bytes, list[str]]:
    mapping = alkatresz_map or load_alkatresz_map()
    missing: list[str] = []
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\n")

    for row in rows:
        kod = row.get("kod", "")
        mennyiseg = row.get("db", "")
        resolved = mapping.get(kod)
        if not resolved:
            fallback_candidates = _procurement_code_fallbacks(kod)
            for fallback_kod in fallback_candidates:
                resolved = mapping.get(fallback_kod)
                if resolved:
                    break

        if not resolved and kod == "NFAY_ANT_PRA_357x197":
            resolved = "NFAY_ANT_PRA_357x197_KA_NO_U_NFA"
        if not resolved:
            missing.append(kod)
            resolved = kod

        writer.writerow([resolved, mennyiseg])

    return buffer.getvalue().encode("utf-8-sig"), sorted(set(missing))


def read_order_rows_from_bytes(data: bytes) -> list[list[object]]:
    _require_workbook_support()

    if zipfile.is_zipfile(io.BytesIO(data)):
        workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    encoding = "utf-8-sig"
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        encoding = "utf-16"
    text = data.decode(encoding)
    sample = text[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.get_dialect("excel")

    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


def _normalize_order_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_order_numeric(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).replace("\u00a0", " ").replace(" ", "").replace(",", ".").strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _numbers_equal(left: object, right: object) -> bool:
    left_num = _normalize_order_numeric(left)
    right_num = _normalize_order_numeric(right)
    if left_num is None and right_num is None:
        return True
    if left_num is None or right_num is None:
        return False
    return left_num == right_num


def _get_column(row: Sequence[object], index: int) -> object:
    if index - 1 < 0 or index - 1 >= len(row):
        return None
    return row[index - 1]


def _build_invoice_index(rows: Iterable[dict[str, str]]) -> dict[str, deque[dict[str, str]]]:
    index: dict[str, deque[dict[str, str]]] = defaultdict(deque)
    for row in rows:
        code = _normalize_order_text(row.get("kod"))
        index[code].append(row)
    return index


def _build_order_index(order_rows: list[list[object]]) -> dict[str, deque[list[object]]]:
    index: dict[str, deque[list[object]]] = defaultdict(deque)
    for row in order_rows[1:]:
        code = left_until_underscore(_get_column(row, 4))
        index[code].append(row)
    return index


def compare_rows(order_rows: list[list[object]], invoice_rows: list[dict[str, str]], invoice_header: list[str]) -> tuple[list[list[object]], list[bool]]:
    invoice_index = _build_invoice_index(invoice_rows)
    output_rows: list[list[object]] = []
    row_matches: list[bool] = []

    if not order_rows:
        return output_rows, row_matches

    header = list(order_rows[0]) + ["processed_kod", "status", "mismatch_details"]
    output_rows.append(header)
    row_matches.append(True)

    for row in order_rows[1:]:
        order_kod = left_until_underscore(_get_column(row, 4))
        order_qty = _get_column(row, 6)
        order_unit = _get_column(row, 9)
        order_net = _get_column(row, 10)

        invoice_queue = invoice_index.get(order_kod)
        invoice_row = invoice_queue.popleft() if invoice_queue else None

        mismatches: list[str] = []
        if invoice_row is None:
            mismatches.append("missing_invoice_row")
        else:
            if _normalize_order_text(order_kod) != _normalize_order_text(invoice_row.get("kod")):
                mismatches.append("kod")
            if not _numbers_equal(order_qty, invoice_row.get("db")):
                mismatches.append("db")
            if not _numbers_equal(order_unit, invoice_row.get("egyseg_ar")):
                mismatches.append("egyseg_ar")
            if not _numbers_equal(order_net, invoice_row.get("netto_ar")):
                mismatches.append("netto_ar")

        status = "OK" if not mismatches else "Mismatch"
        output_rows.append(list(row) + [order_kod, status, ", ".join(mismatches)])
        row_matches.append(not mismatches)

    return output_rows, row_matches


def compare_invoice_rows(order_rows: list[list[object]], invoice_rows: list[dict[str, str]], invoice_header: list[str]) -> tuple[list[list[object]], list[bool]]:
    order_index = _build_order_index(order_rows)
    header = list(invoice_header) + ["status", "mismatch_details"]
    output_rows: list[list[object]] = [header]
    row_matches: list[bool] = [True]

    for invoice_row in invoice_rows:
        invoice_code = _normalize_order_text(invoice_row.get("kod"))
        order_queue = order_index.get(invoice_code)
        order_row = order_queue.popleft() if order_queue else None

        mismatches: list[str] = []
        if order_row is None:
            mismatches.append("missing_order_row")
        else:
            order_qty = _get_column(order_row, 6)
            order_unit = _get_column(order_row, 9)
            order_net = _get_column(order_row, 10)
            if not _numbers_equal(order_qty, invoice_row.get("db")):
                mismatches.append("db")
            if not _numbers_equal(order_unit, invoice_row.get("egyseg_ar")):
                mismatches.append("egyseg_ar")
            if not _numbers_equal(order_net, invoice_row.get("netto_ar")):
                mismatches.append("netto_ar")

        status = "OK" if not mismatches else "Mismatch"
        row_values = [invoice_row.get(name, "") for name in invoice_header]
        output_rows.append(row_values + [status, ", ".join(mismatches)])
        row_matches.append(not mismatches)

    return output_rows, row_matches


def write_report(rows: list[list[object]], row_matches: list[bool], sheet_name: str, workbook: object | None = None):
    _require_workbook_support()

    if workbook is None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    for row in rows:
        sheet.append(row)

    max_col = sheet.max_column
    for idx, matches in enumerate(row_matches, start=1):
        if idx == 1:
            continue
        fill = GREEN_FILL if matches else RED_FILL
        for col in range(1, max_col + 1):
            sheet.cell(row=idx, column=col).fill = fill

    return workbook


def build_compare_workbook(order_rows: list[list[object]], invoice_rows: list[dict[str, str]]) -> bytes:
    invoice_header = list(FIELDNAMES)
    order_output_rows, order_row_matches = compare_rows(order_rows, invoice_rows, invoice_header)
    invoice_output_rows, invoice_row_matches = compare_invoice_rows(order_rows, invoice_rows, invoice_header)

    workbook = write_report(order_output_rows, order_row_matches, sheet_name="Order_to_Invoice")
    workbook = write_report(invoice_output_rows, invoice_row_matches, sheet_name="Invoice_to_Order", workbook=workbook)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_nettfront_artifacts(pdf_bytes: bytes, order_bytes: bytes | None = None):
    procurement = build_procurement_artifacts(pdf_bytes)
    if order_bytes:
        return {
            "procurement": procurement,
            "compare": build_compare_artifacts(pdf_bytes, order_bytes),
        }

    return NettfrontArtifacts(
        invoice_rows=procurement.invoice_rows,
        invoice_csv=procurement.invoice_csv,
        procurement_csv=procurement.procurement_csv,
        compare_workbook=None,
        missing_codes=procurement.missing_codes,
        order_uploaded=False,
        order_row_count=0,
    )


def create_bundle_zip(job_dir: Path, include_compare: bool) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.write(job_dir / "invoice-output.csv", "invoice-output.csv")
        archive.write(job_dir / "rendeles_sima.csv", "rendeles_sima.csv")
        archive.write(job_dir / "metadata.json", "metadata.json")
        if include_compare and (job_dir / "compare-output.xlsx").exists():
            archive.write(job_dir / "compare-output.xlsx", "compare-output.xlsx")
    return buffer.getvalue()


def build_procurement_artifacts(pdf_bytes: bytes, alkatresz_map: dict[str, str] | None = None) -> ProcurementArtifacts:
    invoice_rows = build_invoice_rows(pdf_bytes)
    invoice_csv = build_invoice_csv(invoice_rows)
    procurement_csv, missing_codes = build_procurement_csv(invoice_rows, alkatresz_map=alkatresz_map)

    return ProcurementArtifacts(
        invoice_rows=invoice_rows,
        invoice_csv=invoice_csv,
        procurement_csv=procurement_csv,
        missing_codes=missing_codes,
    )


def build_compare_artifacts(pdf_bytes: bytes, order_bytes: bytes) -> CompareArtifacts:
    invoice_rows = build_invoice_rows(pdf_bytes)
    invoice_csv = build_invoice_csv(invoice_rows)
    order_rows = read_order_rows_from_bytes(order_bytes)
    if not order_rows:
        raise ValueError("A feltöltött megrendelés fájl üres.")

    compare_workbook = build_compare_workbook(order_rows, invoice_rows)
    return CompareArtifacts(
        invoice_rows=invoice_rows,
        invoice_csv=invoice_csv,
        compare_workbook=compare_workbook,
        order_row_count=max(len(order_rows) - 1, 0),
    )


def create_bundle_archive(job_dir: Path, file_names: Sequence[str]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for file_name in file_names:
            file_path = job_dir / file_name
            if file_path.exists():
                archive.write(file_path, file_name)
    return buffer.getvalue()


def _find_procurement_program() -> Path | None:
    configured = os.getenv(DEFAULT_PROCUREMENT_ENV, "").strip()
    if not configured:
        configured = get_procurement_program_path()
    if not configured:
        return None
    candidate = Path(configured)
    if candidate.exists():
        return candidate
    return None


def _read_settings() -> dict:
    if not SETTINGS_FILE.exists():
        return {}
    try:
        payload = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _write_settings(payload: dict) -> None:
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    SETTINGS_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def get_procurement_program_path() -> str:
    configured = os.getenv(DEFAULT_PROCUREMENT_ENV, "").strip()
    if configured:
        candidate = Path(configured)
        if candidate.exists():
            return str(candidate)
        return ""

    configured = str(_read_settings().get("procurement_app_path", "")).strip()
    if not configured:
        return ""

    candidate = Path(configured)
    return str(candidate) if candidate.exists() else ""


def save_procurement_program_path(raw_path: str) -> str:
    candidate = Path(raw_path.strip().strip('"'))
    if not candidate.exists() or not candidate.is_file():
        raise ValueError("A megadott programfájl nem található.")
    if candidate.suffix.lower() != ".exe":
        raise ValueError("Csak .exe fájl menthető el beszerzési alkalmazásként.")

    payload = _read_settings()
    payload["procurement_app_path"] = str(candidate)
    _write_settings(payload)
    return str(candidate)


def _build_runtime_import_script(csv_path: Path, procurement_program: Path) -> str:
    csv_literal = str(csv_path).replace("'", "''")
    program_literal = str(procurement_program).replace("'", "''")
    return f"""$ErrorActionPreference = "Stop"
Add-Type -AssemblyName System.Windows.Forms
Add-Type @"
using System;
using System.Runtime.InteropServices;
public static class DivianKeyboardState {{
    [DllImport("user32.dll")]
    public static extern short GetAsyncKeyState(int vKey);
}}
"@

$csvPath = '{csv_literal}'
$programPath = '{program_literal}'
$launchDelayMs = 2600
$stepDelayMs = 220
$insertDelayMs = 2800
$hotkeyPollDelayMs = 90
$postHotkeyDelayMs = 420

$wshell = New-Object -ComObject WScript.Shell
$process = Start-Process -FilePath $programPath -PassThru
Start-Sleep -Milliseconds $launchDelayMs

$activated = $false
for ($attempt = 0; $attempt -lt 25; $attempt++) {{
    try {{
        if ($wshell.AppActivate($process.Id)) {{
            $activated = $true
            break
        }}
    }} catch {{
    }}
    Start-Sleep -Milliseconds 250
}}

if (-not $activated) {{
    throw "Nem sikerült aktiválni a beszerzési program ablakát."
}}

$lastHotkeyState = $false
while ($true) {{
    if ($process.HasExited) {{
        throw "A beszerzési program bezáródott az import előtt."
    }}

    $shiftDown = ([DivianKeyboardState]::GetAsyncKeyState(0x10) -band 0x8000) -ne 0
    $spaceDown = ([DivianKeyboardState]::GetAsyncKeyState(0x20) -band 0x8000) -ne 0
    $hotkeyDown = $shiftDown -and $spaceDown
    if ($hotkeyDown -and -not $lastHotkeyState) {{
        break
    }}

    $lastHotkeyState = $hotkeyDown
    Start-Sleep -Milliseconds $hotkeyPollDelayMs
}}

while (([DivianKeyboardState]::GetAsyncKeyState(0x10) -band 0x8000) -ne 0 -or ([DivianKeyboardState]::GetAsyncKeyState(0x20) -band 0x8000) -ne 0) {{
    Start-Sleep -Milliseconds 40
}}
Start-Sleep -Milliseconds $postHotkeyDelayMs

$rows = Get-Content -Path $csvPath -Encoding UTF8 | Where-Object {{ $_.Trim() -ne "" }}
foreach ($row in $rows) {{
    $parts = $row -split ';'
    if ($parts.Count -lt 2) {{
        continue
    }}

    $articleCode = $parts[0].Trim()
    $quantity = $parts[1].Trim()
    if ([string]::IsNullOrWhiteSpace($articleCode)) {{
        continue
    }}

    $wshell.AppActivate($process.Id) | Out-Null
    Set-Clipboard -Value $articleCode
    [System.Windows.Forms.SendKeys]::SendWait('{{INSERT}}')
    Start-Sleep -Milliseconds $insertDelayMs
    [System.Windows.Forms.SendKeys]::SendWait('{{TAB}}{{TAB}}')
    Start-Sleep -Milliseconds $stepDelayMs
    [System.Windows.Forms.SendKeys]::SendWait('^v')
    Start-Sleep -Milliseconds $stepDelayMs
    [System.Windows.Forms.SendKeys]::SendWait('{{TAB}}')
    Start-Sleep -Milliseconds $stepDelayMs
    [System.Windows.Forms.SendKeys]::SendWait($quantity)
    Start-Sleep -Milliseconds $stepDelayMs
    [System.Windows.Forms.SendKeys]::SendWait('{{TAB}}')
    Start-Sleep -Milliseconds 140
    [System.Windows.Forms.SendKeys]::SendWait('{{ENTER}}')
    Start-Sleep -Milliseconds $stepDelayMs
}}
"""


def launch_procurement_flow(job_dir: Path) -> tuple[bool, list[str]]:
    job_dir.mkdir(parents=True, exist_ok=True)
    csv_path = job_dir / "rendeles_sima.csv"
    if not csv_path.exists():
        return False, ["A beszerzési CSV nem található."]

    if os.name != "nt":
        return False, ["Az automatikus import jelenleg Windows alatt érhető el."]

    messages: list[str] = []

    procurement_program = _find_procurement_program()
    if procurement_program is None:
        messages.append(
            "A beszerzési program ezen a gépen nem található. "
            f"Állítsd be a {DEFAULT_PROCUREMENT_ENV} környezeti változót."
        )
        return False, messages

    script_path = job_dir / "nettfront-import.ps1"
    script_path.write_text(_build_runtime_import_script(csv_path, procurement_program), encoding="utf-8")
    subprocess.Popen(
        [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
        ],
        cwd=str(job_dir),
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )
    messages.append("A beszerzési program elindult. Az import a Shift + Space megnyomására indul.")
    return True, messages


def _find_autohotkey_executable() -> Path | None:
    candidates = [
        shutil.which("AutoHotkey64.exe"),
        shutil.which("AutoHotkey32.exe"),
        shutil.which("AutoHotkey.exe"),
        r"C:\Program Files\AutoHotkey\v2\AutoHotkey64.exe",
        r"C:\Program Files\AutoHotkey\v2\AutoHotkey32.exe",
        r"C:\Program Files\AutoHotkey\v2\AutoHotkey.exe",
        r"C:\Program Files\AutoHotkey\AutoHotkey64.exe",
        r"C:\Program Files\AutoHotkey\AutoHotkey32.exe",
        r"C:\Program Files\AutoHotkey\AutoHotkey.exe",
        r"C:\Program Files (x86)\AutoHotkey\AutoHotkeyU64.exe",
        r"C:\Program Files (x86)\AutoHotkey\AutoHotkeyU32.exe",
        r"C:\Program Files (x86)\AutoHotkey\AutoHotkey.exe",
    ]
    for candidate in candidates:
        if not candidate:
            continue
        candidate_path = Path(candidate)
        if candidate_path.exists():
            return candidate_path
    return None


def _build_runtime_ahk_script(csv_path: Path) -> str:
    csv_literal = str(csv_path).replace("\\", "\\\\")
    return f"""; AutoHotkey v2.0+
#SingleInstance Force
Persistent

sleeptime := 300
doublesleeptime := 3000
csvPath := "{csv_literal}"

+Space::
{{
    global csvPath, sleeptime, doublesleeptime

    try content := FileRead(csvPath, "UTF-8")
    catch {{
        MsgBox "Hiba: Nem sikerült beolvasni a fájlt: " csvPath
        return
    }}

    lines := StrSplit(content, "`n", "`r")
    for _, line in lines
    {{
        line := Trim(line)
        if (line = "")
            continue

        parts := StrSplit(line, ";")
        if parts.Length < 2
            continue

        cikkszam := Trim(parts[1])
        mennyiseg := Trim(parts[2])
        if (cikkszam = "")
            continue

        A_Clipboard := cikkszam
        Send "{{Insert}}"
        Sleep doublesleeptime

        Send "{{Tab 2}}"
        Sleep sleeptime

        Send "^v"
        Sleep sleeptime

        Send "{{Tab}}"
        Sleep sleeptime

        A_Clipboard := mennyiseg
        Send "^v"
        Sleep sleeptime

        Send "{{Tab}}"
        Send "{{Enter}}"
        Sleep sleeptime
    }}

    MsgBox "Az importálás befejeződött."
}}

+Enter::ExitApp
"""


def _build_runtime_import_script(csv_path: Path) -> str:  # type: ignore[no-redef]
    csv_literal = str(csv_path).replace("'", "''")
    return f"""$ErrorActionPreference = "Stop"
Add-Type -AssemblyName System.Windows.Forms
Add-Type @"
using System;
using System.Runtime.InteropServices;
public static class DivianKeyboardState {{
    [DllImport("user32.dll")]
    public static extern short GetAsyncKeyState(int vKey);
}}
"@

$csvPath = '{csv_literal}'
$stepDelayMs = 220
$insertDelayMs = 2800
$hotkeyPollDelayMs = 90
$postHotkeyDelayMs = 420

$lastHotkeyState = $false
while ($true) {{
    $shiftDown = ([DivianKeyboardState]::GetAsyncKeyState(0x10) -band 0x8000) -ne 0
    $spaceDown = ([DivianKeyboardState]::GetAsyncKeyState(0x20) -band 0x8000) -ne 0
    $enterDown = ([DivianKeyboardState]::GetAsyncKeyState(0x0D) -band 0x8000) -ne 0
    if ($shiftDown -and $enterDown) {{
        exit 0
    }}

    $hotkeyDown = $shiftDown -and $spaceDown
    if ($hotkeyDown -and -not $lastHotkeyState) {{
        break
    }}

    $lastHotkeyState = $hotkeyDown
    Start-Sleep -Milliseconds $hotkeyPollDelayMs
}}

while (([DivianKeyboardState]::GetAsyncKeyState(0x10) -band 0x8000) -ne 0 -or ([DivianKeyboardState]::GetAsyncKeyState(0x20) -band 0x8000) -ne 0) {{
    Start-Sleep -Milliseconds 40
}}
Start-Sleep -Milliseconds $postHotkeyDelayMs

$rows = Get-Content -Path $csvPath -Encoding UTF8 | Where-Object {{ $_.Trim() -ne "" }}
foreach ($row in $rows) {{
    $parts = $row -split ';'
    if ($parts.Count -lt 2) {{
        continue
    }}

    $articleCode = $parts[0].Trim()
    $quantity = $parts[1].Trim()
    if ([string]::IsNullOrWhiteSpace($articleCode)) {{
        continue
    }}

    Set-Clipboard -Value $articleCode
    [System.Windows.Forms.SendKeys]::SendWait('{{INSERT}}')
    Start-Sleep -Milliseconds $insertDelayMs
    [System.Windows.Forms.SendKeys]::SendWait('{{TAB}}{{TAB}}')
    Start-Sleep -Milliseconds $stepDelayMs
    [System.Windows.Forms.SendKeys]::SendWait('^v')
    Start-Sleep -Milliseconds $stepDelayMs
    [System.Windows.Forms.SendKeys]::SendWait('{{TAB}}')
    Start-Sleep -Milliseconds $stepDelayMs
    Set-Clipboard -Value $quantity
    [System.Windows.Forms.SendKeys]::SendWait('^v')
    Start-Sleep -Milliseconds $stepDelayMs
    [System.Windows.Forms.SendKeys]::SendWait('{{TAB}}')
    Start-Sleep -Milliseconds 140
    [System.Windows.Forms.SendKeys]::SendWait('{{ENTER}}')
    Start-Sleep -Milliseconds $stepDelayMs
}}
"""


def launch_procurement_flow(job_dir: Path) -> tuple[bool, list[str]]:  # type: ignore[no-redef]
    job_dir.mkdir(parents=True, exist_ok=True)
    csv_path = job_dir / "rendeles_sima.csv"
    if not csv_path.exists():
        return False, ["A Beszerzés fájl nem található."]

    if os.name != "nt":
        return False, ["Az automatikus import jelenleg Windows alatt érhető el."]

    ahk_path = job_dir / "nettfront-import.ahk"
    ahk_path.write_text(_build_runtime_ahk_script(csv_path), encoding="utf-8")

    script_path = job_dir / "nettfront-import.ps1"
    script_path.write_text(_build_runtime_import_script(csv_path), encoding="utf-8")

    autohotkey = _find_autohotkey_executable()
    if autohotkey is not None:
        subprocess.Popen(
            [str(autohotkey), str(ahk_path)],
            cwd=str(job_dir),
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        return True, ["Az import-segéd elindult. Nyisd meg a beszerzési ablakot, majd nyomd meg a Shift + Space-t."]

    subprocess.Popen(
        [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
        ],
        cwd=str(job_dir),
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )
    return True, ["Az import-segéd elindult. Nyisd meg a beszerzési ablakot, majd nyomd meg a Shift + Space-t."]
