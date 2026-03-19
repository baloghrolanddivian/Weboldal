from __future__ import annotations

import io
import math
import re
import unicodedata as ud
import warnings
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl as ox
from openpyxl.utils import get_column_letter


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
)

ANT_VARIANT_RULES = (
    ("KAF", "KAFS", "KAFU"),
    ("PRA", "PRAS", "PRAU"),
)

SUGGESTION_HEADERS = [
    "Alkatr.-leírás",
    "Rend.áll.rakt.készl. ME",
    "Rend.áll",
    "Visszaigazolt igény",
    "Nyitott beszerzés",
    "Biztonsági készlet",
    "Tárolh.menny.",
    "Rendelendő mennyiseg",
]


@dataclass
class NettfrontOrderRow:
    row_id: str
    part_number: str
    description: str
    stock_unit: float | int | str | None
    current_stock: float | int | str | None
    confirmed_demand: float | int | str | None
    open_procurement: float | int | str | None
    safe_stock: float | int | str | None
    capacity: float | int | str | None
    order_qty: float
    color: str
    length: float
    width: float
    is_super_matt: bool = False


@dataclass
class NettfrontOrderBuildResult:
    rows: list[NettfrontOrderRow]
    merged_variant_count: int
    filtered_stock_count: int
    added_super_matt_count: int
    total_m2: float
    avg_row_count: int
    suggestion_workbook: bytes


def norm(value) -> str:
    if value is None:
        return ""
    text = "".join(c for c in ud.normalize("NFD", str(value)) if ud.category(c) != "Mn")
    return "".join(c for c in text.lower() if c.isalnum())


def _safe_number(value) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return 0.0
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _display_number(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _workbook_from_bytes(payload: bytes):
    return ox.load_workbook(io.BytesIO(payload), data_only=True)


def load_avg_by_part_from_bytes(payload: bytes) -> dict[str, int]:
    wb = _workbook_from_bytes(payload)
    ws = wb.active
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    norm_headers = {norm(name): idx for idx, name in enumerate(headers)}

    if "alkatreszszam" in norm_headers and "atlagosmennyiseg12hofelfele" in norm_headers:
        idx_part = norm_headers["alkatreszszam"]
        idx_avg = norm_headers["atlagosmennyiseg12hofelfele"]
        avg_by_part: dict[str, int] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            part = row[idx_part]
            avg = row[idx_avg]
            if part in (None, "") or avg in (None, ""):
                continue
            avg_by_part[str(part)] = int(math.ceil(_safe_number(avg)))
        return avg_by_part

    required = ["mennyiseg", "ev", "honap", "alkatrszam"]
    missing = [name for name in required if name not in norm_headers]
    if missing:
        raise ValueError(f"Hiányzó oszlop(ok) az átlagfájlban: {', '.join(missing)}")

    idx_qty = norm_headers["mennyiseg"]
    idx_year = norm_headers["ev"]
    idx_month = norm_headers["honap"]
    idx_part = norm_headers["alkatrszam"]

    rows: list[tuple[int, str, float]] = []
    max_month_index: int | None = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        year = row[idx_year]
        month = row[idx_month]
        part = row[idx_part]
        qty = row[idx_qty]
        if year is None or month is None or part in (None, "") or qty is None:
            continue
        try:
            month_index = int(year) * 12 + int(month)
        except (TypeError, ValueError):
            continue
        if max_month_index is None or month_index > max_month_index:
            max_month_index = month_index
        rows.append((month_index, str(part), _safe_number(qty)))

    if max_month_index is None:
        raise ValueError("Nem találtam használható dátumos sort az átlagfájlban.")

    min_month_index = max_month_index - 11
    sums = defaultdict(float)
    counts = defaultdict(int)
    for month_index, part, qty in rows:
        if month_index < min_month_index:
            continue
        sums[part] += qty
        counts[part] += 1

    return {part: int(math.ceil(total / counts[part])) for part, total in sums.items() if counts[part]}


def load_avg_by_part(avg_bytes: bytes | None, default_avg_path: Path | None = None) -> dict[str, int]:
    if avg_bytes:
        return load_avg_by_part_from_bytes(avg_bytes)
    if default_avg_path and default_avg_path.exists():
        return load_avg_by_part_from_bytes(default_avg_path.read_bytes())
    raise ValueError("Nem található átlagfájl a rendelési javaslathoz.")


def _load_stock_headers_and_rows(stock_bytes: bytes) -> tuple[list, list[tuple]]:
    wb = _workbook_from_bytes(stock_bytes)
    ws = wb.active
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = [tuple(row) for row in ws.iter_rows(min_row=2, values_only=True)]
    return headers, rows


def is_super_matt(*values) -> bool:
    for value in values:
        if value is None:
            continue
        text = str(value).lower()
        if "sm." in text:
            return True
        cleaned = "".join(c if c.isalnum() else " " for c in text)
        if "sm" in {token for token in cleaned.split() if token}:
            return True
    return False


def is_allowed_legacy_matt_exception(desc) -> bool:
    desc_norm = norm(desc)
    return "antonia" in desc_norm and ("kasmir" in desc_norm or "provance" in desc_norm)


def is_custom_y_item(part=None, desc=None) -> bool:
    part_text = "" if part is None else str(part).strip().upper()
    desc_text = "" if desc is None else str(desc).strip().lower()
    return part_text.startswith("NFAY_") or desc_text.startswith("egyedi f.")


def is_legacy_matt_to_exclude(desc, part=None) -> bool:
    if desc in (None, ""):
        return False
    if is_custom_y_item(part, desc):
        return False
    if is_super_matt(desc):
        return False
    if is_allowed_legacy_matt_exception(desc):
        return False
    text = str(desc).lower()
    cleaned = "".join(c if c.isalnum() else " " for c in text)
    return "matt" in {token for token in cleaned.split() if token}


def collect_replaced_ant_variant_parts(parts) -> set[str]:
    families: dict[tuple[str, str, str, str, str], set[str]] = {}
    for part in parts:
        part_str = str(part)
        for a_tag, b_tag, u_tag in ANT_VARIANT_RULES:
            match = re.match(rf"^(NFA[^_]*_ANT)_({a_tag}|{b_tag}|{u_tag})_(.+)$", part_str)
            if not match:
                continue
            key = (match.group(1), match.group(3), a_tag, b_tag, u_tag)
            families.setdefault(key, set()).add(match.group(2))
            break

    replaced_parts: set[str] = set()
    for (base, suffix, a_tag, b_tag, u_tag), tags in families.items():
        if u_tag in tags or (a_tag in tags and b_tag in tags):
            if a_tag in tags:
                replaced_parts.add(f"{base}_{a_tag}_{suffix}")
            if b_tag in tags:
                replaced_parts.add(f"{base}_{b_tag}_{suffix}")
    return replaced_parts


def calculate_order_qty(current: float, capacity: float, desc=None) -> int:
    desc_text = "" if desc is None else str(desc).strip().lower()
    if desc_text.startswith("egyedi f.") and current < 0:
        return int(math.ceil(-current))

    diff = capacity - current
    if diff <= 0:
        return 0
    return int(diff // 5) * 5


def _merge_ant_variant_rows(headers: list, rows: list[tuple]) -> tuple[list[tuple], int]:
    norm_headers = {norm(name): idx for idx, name in enumerate(headers)}
    idx_part = norm_headers.get("alkatrszam", 0)
    sum_indices = [2, 3, 4, 5, 6]
    rows_by_part = {
        str(row[idx_part]): list(row)
        for row in rows
        if idx_part < len(row) and row[idx_part] not in (None, "")
    }
    merged_rows = list(rows)
    merged = 0

    for part, row in list(rows_by_part.items()):
        matched = None
        for a_tag, b_tag, u_tag in ANT_VARIANT_RULES:
            hit = re.match(rf"^(NFA[^_]*_ANT)_{a_tag}_(.+)$", part)
            if hit:
                matched = (hit.group(1), hit.group(2), a_tag, b_tag, u_tag)
                break
        if matched is None:
            continue

        base, suffix, a_tag, b_tag, u_tag = matched
        pair_part = f"{base}_{b_tag}_{suffix}"
        merged_part = f"{base}_{u_tag}_{suffix}"
        if pair_part not in rows_by_part or merged_part in rows_by_part:
            continue

        pair_row = rows_by_part[pair_part]
        new_row = list(row)
        new_row[idx_part] = merged_part
        for idx in sum_indices:
            if idx >= len(new_row) or idx >= len(pair_row):
                continue
            new_row[idx] = _safe_number(new_row[idx]) + _safe_number(pair_row[idx])

        rows_by_part[merged_part] = new_row
        merged_rows.append(tuple(new_row))
        merged += 1

    return merged_rows, merged


def _build_row_id(part_number: str, description: str, index: int) -> str:
    seed = f"{index}:{part_number}:{description}"
    return re.sub(r"[^a-z0-9]+", "-", norm(seed)).strip("-") or f"row-{index}"


def _row_to_suggestion(
    row: tuple,
    headers: list,
    order_qty: int,
    index: int,
    *,
    is_super_matt_row: bool = False,
) -> NettfrontOrderRow:
    norm_headers = {norm(name): idx for idx, name in enumerate(headers)}
    idx_part = norm_headers.get("alkatrszam", 0)
    idx_desc = norm_headers.get("alkatrleiras", 1)
    idx_stock_unit = norm_headers.get("rendallraktkeszlme", 2)
    idx_current = norm_headers.get("rendall", 3)
    idx_confirmed = norm_headers.get("visszaigazoltigeny", 5)
    idx_open_proc = norm_headers.get("nyitottbeszerzes", 6)
    idx_safe = norm_headers.get("biztonsagikeszlet", 7)
    idx_capacity = norm_headers.get("tarolhmenny", 8)
    idx_length = norm_headers.get("hossz", 9)
    idx_width = norm_headers.get("szelesseg", 10)
    idx_color = norm_headers.get("szinnettfrdesc", 11)

    part_number = "" if idx_part is None or idx_part >= len(row) or row[idx_part] is None else str(row[idx_part])
    description = "" if idx_desc is None or idx_desc >= len(row) or row[idx_desc] is None else str(row[idx_desc])

    return NettfrontOrderRow(
        row_id=_build_row_id(part_number, description, index),
        part_number=part_number,
        description=description,
        stock_unit=row[idx_stock_unit] if idx_stock_unit is not None and idx_stock_unit < len(row) else None,
        current_stock=row[idx_current] if idx_current is not None and idx_current < len(row) else None,
        confirmed_demand=row[idx_confirmed] if idx_confirmed is not None and idx_confirmed < len(row) else None,
        open_procurement=row[idx_open_proc] if idx_open_proc is not None and idx_open_proc < len(row) else None,
        safe_stock=row[idx_safe] if idx_safe is not None and idx_safe < len(row) else None,
        capacity=row[idx_capacity] if idx_capacity is not None and idx_capacity < len(row) else None,
        order_qty=float(order_qty),
        color="" if idx_color is None or idx_color >= len(row) or row[idx_color] is None else str(row[idx_color]),
        length=_safe_number(row[idx_length]) if idx_length is not None and idx_length < len(row) else 0.0,
        width=_safe_number(row[idx_width]) if idx_width is not None and idx_width < len(row) else 0.0,
        is_super_matt=is_super_matt_row,
    )


def rows_to_suggestion_workbook(rows: list[NettfrontOrderRow]) -> bytes:
    wb = ox.Workbook()
    ws = wb.active
    ws.title = "rendeles"
    ws.append(SUGGESTION_HEADERS)
    for row in rows:
        ws.append(
            [
                row.description,
                _display_number(row.stock_unit),
                _display_number(row.current_stock),
                _display_number(row.confirmed_demand),
                _display_number(row.open_procurement),
                _display_number(row.safe_stock),
                _display_number(row.capacity),
                int(row.order_qty) if float(row.order_qty).is_integer() else row.order_qty,
            ]
        )

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for current in cell:
                value = "" if current.value is None else str(current.value)
                if len(value) > max_len:
                    max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 42)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def rows_to_approved_workbook(rows: list[NettfrontOrderRow], title: str) -> bytes:
    wb = ox.Workbook()
    ws = wb.active
    ws.title = "jovahagyott_rendeles"
    ws.append([title, None, None])

    export_rows = [
        row for row in rows
        if row.description.strip() and _safe_number(row.order_qty) > 0
    ]
    export_rows.sort(key=lambda item: item.description.lower(), reverse=True)

    for row in export_rows:
        qty = _safe_number(row.order_qty)
        qty_out = int(qty) if float(qty).is_integer() else qty
        ws.append([row.description, qty_out, row.color])

    for col_idx in range(1, 4):
        max_len = 0
        for column in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def calc_total_m2_from_rows(rows: list[NettfrontOrderRow]) -> float:
    total = 0.0
    for row in rows:
        qty = _safe_number(row.order_qty)
        if qty <= 0:
            continue
        total += (row.length * row.width / 1_000_000.0) * qty
    return total


def build_order_suggestions(
    stock_bytes: bytes,
    avg_bytes: bytes | None = None,
    *,
    default_avg_path: Path | None = None,
) -> NettfrontOrderBuildResult:
    headers, stock_rows = _load_stock_headers_and_rows(stock_bytes)
    stock_rows, merged_variant_count = _merge_ant_variant_rows(headers, stock_rows)
    avg_by_part = load_avg_by_part(avg_bytes, default_avg_path)

    norm_headers = {norm(name): idx for idx, name in enumerate(headers)}
    idx_part = norm_headers.get("alkatrszam", 0)
    idx_desc = norm_headers.get("alkatrleiras", 1)
    idx_stock_unit = norm_headers.get("rendallraktkeszlme")
    idx_current = norm_headers.get("rendall")
    idx_safe = norm_headers.get("biztonsagikeszlet")
    idx_capacity = norm_headers.get("tarolhmenny")

    if idx_current is None or idx_safe is None or idx_capacity is None:
        raise ValueError("A raktárfájlban hiányzik a Rend.áll, Biztonsági készlet vagy Tárolh.menny. oszlop.")

    replaced_parts = collect_replaced_ant_variant_parts(
        row[idx_part]
        for row in stock_rows
        if idx_part < len(row) and row[idx_part] not in (None, "")
    )

    filtered_rows: list[tuple] = []
    order_by_part: dict[str, int] = {}
    for row in stock_rows:
        part = row[idx_part] if idx_part < len(row) else None
        desc = row[idx_desc] if idx_desc < len(row) else None
        stock_unit = row[idx_stock_unit] if idx_stock_unit is not None and idx_stock_unit < len(row) else None
        current_value = row[idx_current] if idx_current < len(row) else None
        safe_value = row[idx_safe] if idx_safe < len(row) else None

        if part in replaced_parts:
            continue
        if desc is not None and is_super_matt(desc):
            continue
        if is_legacy_matt_to_exclude(desc, part):
            continue
        if (stock_unit is None or str(stock_unit).strip() == "") and (current_value is None or str(current_value).strip() == ""):
            continue

        current = _safe_number(current_value)
        safe_stock = _safe_number(safe_value)
        if current < safe_stock:
            filtered_rows.append(row)
            order_by_part[str(part)] = calculate_order_qty(current, _safe_number(row[idx_capacity]), desc)

    suggestion_rows: list[NettfrontOrderRow] = []
    existing_descs: set[str] = set()
    row_index = 1
    for row in stock_rows:
        part = row[idx_part] if idx_part < len(row) else None
        if part in (None, ""):
            continue
        order_qty = order_by_part.get(str(part))
        if order_qty is None:
            continue
        suggestion = _row_to_suggestion(row, headers, order_qty, row_index)
        suggestion_rows.append(suggestion)
        existing_descs.add(norm(suggestion.description))
        row_index += 1

    added_super_matt_count = 0
    for row in stock_rows:
        part = row[idx_part] if idx_part < len(row) else None
        desc = row[idx_desc] if idx_desc < len(row) else None
        if part in (None, "") or desc in (None, ""):
            continue
        if str(part) in replaced_parts:
            continue
        if not is_super_matt(desc):
            continue
        if norm(desc) in existing_descs:
            continue

        current = _safe_number(row[idx_current])
        safe_stock = _safe_number(row[idx_safe])
        if current >= safe_stock:
            continue

        order_qty = calculate_order_qty(current, _safe_number(row[idx_capacity]), desc)
        suggestion = _row_to_suggestion(row, headers, order_qty, row_index, is_super_matt_row=True)
        suggestion_rows.append(suggestion)
        existing_descs.add(norm(suggestion.description))
        added_super_matt_count += 1
        row_index += 1

    suggestion_workbook = rows_to_suggestion_workbook(suggestion_rows)
    return NettfrontOrderBuildResult(
        rows=suggestion_rows,
        merged_variant_count=merged_variant_count,
        filtered_stock_count=len(filtered_rows),
        added_super_matt_count=added_super_matt_count,
        total_m2=calc_total_m2_from_rows(suggestion_rows),
        avg_row_count=len(avg_by_part),
        suggestion_workbook=suggestion_workbook,
    )
