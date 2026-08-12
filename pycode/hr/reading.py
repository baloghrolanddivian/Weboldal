"""Read HR spreadsheets without writing the uploaded contents to disk."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None

HR_COLUMNS = (
    "name", "vat", "address", "job", "birthname", "birthplace", "birthday",
    "momname", "taj", "entry", "payment", "stayaddress", "email", "phone",
)


def _value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y.%m.%d.")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _xlsx_rows(file_data: bytes) -> list[tuple[object, ...]]:
    if load_workbook is None:
        raise RuntimeError("Az XLSX feldolgozásához az openpyxl csomag szükséges.")
    workbook = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _xls_rows(file_data: bytes) -> list[tuple[object, ...]]:
    if xlrd is None:
        raise RuntimeError("Az XLS feldolgozásához az xlrd csomag szükséges.")
    workbook = xlrd.open_workbook(file_contents=file_data, on_demand=True)
    try:
        sheet = workbook.sheet_by_index(0)
        rows = []
        for row_index in range(sheet.nrows):
            values = []
            for cell in sheet.row(row_index):
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, workbook.datemode)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(value)
                elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    value = None
                values.append(value)
            rows.append(tuple(values))
        return rows
    finally:
        workbook.release_resources()


def read_people(file_data: bytes) -> list[dict[str, str]]:
    """Parse XLS/XLSX rows from memory; uploaded contents are never persisted."""
    if file_data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        rows = _xls_rows(file_data)
    elif file_data.startswith(b"PK"):
        rows = _xlsx_rows(file_data)
    else:
        raise ValueError("Nem támogatott Excel-formátum. Használj .xls, .xlsx vagy .xlsm fájlt.")
    people = []
    # Row 1 is the spreadsheet header/description row and must never become an employee.
    for row in rows[1:]:
        values = [_value(row[index] if index < len(row) else "") for index in range(len(HR_COLUMNS))]
        if not any(values):
            continue
        people.append(dict(zip(HR_COLUMNS, values)))
    if not people:
        raise ValueError("Az Excel nem tartalmaz feldolgozható személy sort.")
    return people
