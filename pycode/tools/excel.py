"""Compatibility helpers for uploaded Microsoft Excel workbooks."""

from __future__ import annotations

import io

try:
    import xlrd
except Exception:  # pragma: no cover - optional dependency handling
    xlrd = None

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover - optional dependency handling
    Workbook = None


LEGACY_XLS_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def is_legacy_xls(payload: bytes) -> bool:
    """Return whether payload is an OLE-based Excel 97-2003 workbook."""
    return payload.startswith(LEGACY_XLS_SIGNATURE)


def normalize_excel_payload(payload: bytes) -> bytes:
    """Convert a legacy XLS workbook to XLSX bytes for openpyxl consumers."""
    if not is_legacy_xls(payload):
        return payload
    if xlrd is None:
        raise RuntimeError("Az XLS feldolgozásához az xlrd csomag szükséges.")
    if Workbook is None:
        raise RuntimeError("Az Excel feldolgozásához az openpyxl csomag szükséges.")

    source = xlrd.open_workbook(file_contents=payload, on_demand=True)
    target = Workbook()
    target.remove(target.active)
    try:
        for source_sheet in source.sheets():
            target_sheet = target.create_sheet(source_sheet.name)
            for row_index in range(source_sheet.nrows):
                for column_index, cell in enumerate(source_sheet.row(row_index), start=1):
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate_as_datetime(value, source.datemode)
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        value = bool(value)
                    elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        value = None
                    elif cell.ctype == xlrd.XL_CELL_ERROR:
                        value = xlrd.error_text_from_code.get(value, "#ERROR!")
                    target_sheet.cell(row=row_index + 1, column=column_index, value=value)

        output = io.BytesIO()
        target.save(output)
        return output.getvalue()
    finally:
        source.release_resources()
        target.close()
