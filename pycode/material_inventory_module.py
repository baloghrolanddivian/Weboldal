from __future__ import annotations

import csv
import io
import json
import re
import secrets
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None


MATERIAL_INVENTORY_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}


def file_name_allowed(file_name: str) -> bool:
    return Path(file_name or "").suffix.lower() in MATERIAL_INVENTORY_ALLOWED_EXTENSIONS


def read_bytes_if_exists(path: Path) -> bytes | None:
    if not path.exists():
        return None
    return path.read_bytes()


def write_runtime_upload(base_path: Path, file_name: str, payload: bytes) -> Path:
    suffix = Path(file_name or "").suffix.lower() or ".bin"
    target_path = base_path.with_suffix(suffix)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(payload)
    return target_path


def load_session_from_path(path: Path) -> dict | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return payload if isinstance(payload, dict) else None


def save_session_to_path(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_material_inventory_session(file_name: str, payload: bytes) -> dict:
    source_rows = _read_material_rows(file_name, payload)
    rows: list[dict] = []
    for index, item in enumerate(source_rows, start=1):
        if _is_excluded(item.get("exclude")):
            continue
        part_number = _clean_text(item.get("part_number"))
        if not part_number:
            continue
        category = _clean_text(item.get("icg_code")) or "Kategória nélkül"
        rows.append(
            {
                "row_id": _row_id(part_number, category, str(index)),
                "part_number": part_number,
                "description": _clean_text(item.get("description")),
                "book_qty": _clean_number_text(item.get("book_qty")),
                "icg_code": category,
                "input_qty": "",
            }
        )

    rows.sort(key=_row_sort_key)
    return {
        "session_id": secrets.token_hex(6),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "source_name": Path(file_name).name,
        "category_label": "ICG kód",
        "export_prefix": "anyag-raktar",
        "phase": "counting",
        "phase_label": "Számlálás",
        "finalized_at": "",
        "rows": rows,
    }


def build_semifinished_inventory_session(file_name: str, payload: bytes) -> dict:
    return _build_color_inventory_session(file_name, payload, "felkesz-raktar")


def build_semifinished_front_inventory_session(file_name: str, payload: bytes) -> dict:
    return _build_color_inventory_session(file_name, payload, "felkesz-front")


def _build_color_inventory_session(file_name: str, payload: bytes, export_prefix: str) -> dict:
    source_rows = _read_semifinished_rows(file_name, payload)
    rows: list[dict] = []
    for index, item in enumerate(source_rows, start=1):
        if _is_excluded(item.get("exclude")):
            continue
        part_number = _clean_text(item.get("part_number"))
        if not part_number:
            continue
        category = _clean_text(item.get("color_desc")) or _clean_text(item.get("color_code")) or "Szín nélkül"
        rows.append(
            {
                "row_id": _row_id(part_number, category, str(index)),
                "part_number": part_number,
                "description": _clean_text(item.get("description")),
                "book_qty": "",
                "icg_code": category,
                "input_qty": "",
            }
        )

    rows.sort(key=_row_sort_key)
    return {
        "session_id": secrets.token_hex(6),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "source_name": Path(file_name).name,
        "category_label": "Szín",
        "export_prefix": export_prefix,
        "phase": "counting",
        "phase_label": "Számlálás",
        "finalized_at": "",
        "rows": rows,
    }


def update_material_row_input(session: dict, row_id: str, raw_value: str, mode: str = "set") -> tuple[bool, str]:
    row = _find_row(session, row_id)
    if row is None:
        return False, "A kiválasztott anyagsort nem találom."
    if str(session.get("phase", "")).lower() == "finalized":
        return False, "A leltár már le van zárva."

    clean_value = str(raw_value or "").strip()
    clean_mode = str(mode or "set").strip().lower()
    if clean_mode not in {"set", "add", "subtract"}:
        clean_mode = "set"

    if not clean_value:
        if clean_mode != "set":
            return True, ""
        row["input_qty"] = ""
        _touch_session(session)
        return True, ""

    parsed_value = _parse_non_negative_number(clean_value)
    if parsed_value is None:
        return False, "Csak nem negatív szám adható meg."
    if clean_mode == "set":
        next_value = parsed_value
    else:
        current_value = _parse_non_negative_number(row.get("input_qty")) or 0
        next_value = current_value + parsed_value if clean_mode == "add" else current_value - parsed_value
        if next_value < 0:
            return False, "A levonás után nem lehet negatív a darabszám."
    row["input_qty"] = _format_quantity(next_value)
    _touch_session(session)
    return True, ""


def build_material_inventory_view_model(session: dict, selected_category: str = "") -> dict:
    rows = [row for row in session.get("rows", []) if isinstance(row, dict)]
    categories = _category_summaries(rows)
    valid_keys = {item["key"] for item in categories}
    selected = selected_category if selected_category in valid_keys else "all"
    visible_rows = rows if selected == "all" else [row for row in rows if _category_key(row.get("icg_code")) == selected]
    visible_rows = sorted(visible_rows, key=_row_sort_key)
    finalized = str(session.get("phase", "")).lower() == "finalized"
    return {
        "source_name": str(session.get("source_name", "")),
        "updated_at": str(session.get("updated_at", "")),
        "phase_label": str(session.get("phase_label", "Számlálás")),
        "finalized_at": str(session.get("finalized_at", "")),
        "finalized": finalized,
        "categories": categories,
        "selected_category": selected,
        "visible_rows": visible_rows,
        "total_rows": len(rows),
        "counted_rows": sum(1 for row in rows if str(row.get("input_qty", "")).strip()),
        "missing_rows": sum(1 for row in rows if not str(row.get("input_qty", "")).strip()),
        "category_count": max(len(categories) - 1, 0),
    }


def finalize_material_inventory(session: dict, allow_missing: bool = True) -> tuple[bool, str]:
    if str(session.get("phase", "")).lower() == "finalized":
        return True, "Az anyagraktár leltár már le van zárva."
    rows = [row for row in session.get("rows", []) if isinstance(row, dict)]
    missing_count = sum(1 for row in rows if not str(row.get("input_qty", "")).strip())
    if missing_count and not allow_missing:
        return False, f"Még {missing_count} tételnél nincs darabszám."
    for row in rows:
        value = _parse_non_negative_number(row.get("input_qty"))
        row["counted_qty"] = _format_quantity(value if value is not None else 0)
    session["phase"] = "finalized"
    session["phase_label"] = "Lezárva"
    session["finalized_at"] = datetime.now().isoformat(timespec="seconds")
    _touch_session(session)
    return True, "Az anyagraktár leltár lezárva, az exportok elkészültek."


def build_material_inventory_insight_workbook(session: dict) -> tuple[bytes, str, int]:
    if Workbook is None:
        raise RuntimeError("Az Excel exporthoz hianyzik az openpyxl csomag.")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "InSight"
    headers = ["Alkatr.-szám", "Darabszám"]
    sheet.append(headers)
    rows = sorted([row for row in session.get("rows", []) if isinstance(row, dict)], key=_row_sort_key)
    for row in rows:
        sheet.append([row.get("part_number", ""), _excel_quantity(row)])
    _style_sheet(sheet, widths=(28, 16))
    buffer = io.BytesIO()
    workbook.save(buffer)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    prefix = _clean_text(session.get("export_prefix")) or "anyag-raktar"
    return buffer.getvalue(), f"{prefix}-insight-{stamp}.xlsx", len(rows)


def build_material_inventory_summary_workbook(session: dict) -> tuple[bytes, str, int]:
    if Workbook is None:
        raise RuntimeError("Az Excel exporthoz hianyzik az openpyxl csomag.")
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Összesítő"
    category_label = _clean_text(session.get("category_label")) or "ICG kód"
    summary_sheet.append([category_label, "Tételek", "Kitöltött sorok", "Összes darabszám"])
    rows = sorted([row for row in session.get("rows", []) if isinstance(row, dict)], key=_row_sort_key)
    for item in _summary_by_category(rows):
        summary_sheet.append([item["icg_code"], item["row_count"], item["counted_count"], item["total_qty"]])
    _style_sheet(summary_sheet, widths=(26, 16, 18, 18))

    detail_sheet = workbook.create_sheet("Tételek")
    detail_sheet.append(
        [
            category_label,
            "Alkatr.-szám",
            "Alkatr.-leírás",
            "Könyvelési mennyiség",
            "Számolt darabszám",
        ]
    )
    for row in rows:
        detail_sheet.append(
            [
                row.get("icg_code", ""),
                row.get("part_number", ""),
                row.get("description", ""),
                row.get("book_qty", ""),
                _excel_quantity(row),
            ]
        )
    _style_sheet(detail_sheet, widths=(20, 28, 50, 22, 18))

    buffer = io.BytesIO()
    workbook.save(buffer)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    prefix = _clean_text(session.get("export_prefix")) or "anyag-raktar"
    return buffer.getvalue(), f"{prefix}-osszesito-{stamp}.xlsx", len(rows)


def _read_material_rows(file_name: str, payload: bytes) -> list[dict]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(payload)
    return _read_xlsx_rows(payload)


def _read_xlsx_rows(payload: bytes) -> list[dict]:
    if load_workbook is None:
        raise RuntimeError("Az XLSX olvasáshoz hiányzik az openpyxl csomag.")
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("A feltöltött anyagraktár fájl üres.") from exc
    header_map = _header_map(header_row)
    _ensure_required_headers(header_map)
    _apply_material_book_qty_fallback(header_map, header_row)
    rows: list[dict] = []
    for values in rows_iter:
        item = _row_from_values(header_map, values)
        if any(str(value or "").strip() for value in item.values()):
            rows.append(item)
    return rows


def _read_csv_rows(payload: bytes) -> list[dict]:
    text = payload.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    dialect = csv.Sniffer().sniff(sample, delimiters=";,	,")
    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header_row = next(reader)
    except StopIteration as exc:
        raise ValueError("A feltöltött anyagraktár CSV üres.") from exc
    header_map = _header_map(header_row)
    _ensure_required_headers(header_map)
    _apply_material_book_qty_fallback(header_map, header_row)
    rows: list[dict] = []
    for values in reader:
        item = _row_from_values(header_map, values)
        if any(str(value or "").strip() for value in item.values()):
            rows.append(item)
    return rows


def _read_semifinished_rows(file_name: str, payload: bytes) -> list[dict]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix == ".csv":
        return _read_semifinished_csv_rows(payload)
    return _read_semifinished_xlsx_rows(payload)


def _read_semifinished_xlsx_rows(payload: bytes) -> list[dict]:
    if load_workbook is None:
        raise RuntimeError("Az XLSX olvasáshoz hiányzik az openpyxl csomag.")
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("A feltöltött félkész raktár fájl üres.") from exc
    header_map = _header_map(header_row)
    _ensure_semifinished_headers(header_map)
    rows: list[dict] = []
    for values in rows_iter:
        item = _semifinished_row_from_values(header_map, values)
        if any(str(value or "").strip() for value in item.values()):
            rows.append(item)
    return rows


def _read_semifinished_csv_rows(payload: bytes) -> list[dict]:
    text = payload.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    dialect = csv.Sniffer().sniff(sample, delimiters=";,	,")
    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header_row = next(reader)
    except StopIteration as exc:
        raise ValueError("A feltöltött félkész raktár CSV üres.") from exc
    header_map = _header_map(header_row)
    _ensure_semifinished_headers(header_map)
    rows: list[dict] = []
    for values in reader:
        item = _semifinished_row_from_values(header_map, values)
        if any(str(value or "").strip() for value in item.values()):
            rows.append(item)
    return rows


def _header_map(header_row: tuple | list) -> dict[str, int]:
    aliases = {
        "part_number": {"alkatr.-szám", "alkatrész szám", "alkatresz szam", "alkatr-szam", "cikkszám", "cikkszam"},
        "description": {"alkatr.-leírás", "alkatrész leírás", "alkatresz leiras", "leírás", "leiras"},
        "book_qty": {
            "könyvelési mennyiség",
            "konyvelesi mennyiseg",
            "konyvelesi menny",
            "konyveles mennyiseg",
            "könyvelési menny.",
            "konyvelesi menny.",
            "könyvelt mennyiség",
            "konyvelt mennyiseg",
            "konyvelt menny",
            "könyv. menny.",
            "konyv. menny.",
            "konyv mennyiseg",
            "konyv menny",
        },
        "icg_code": {"icg kód", "icg kod", "icg"},
        "color_code": {"szin", "szín", "color", "szinkod", "színkód", "szin kod", "szín kód"},
        "color_desc": {"szin.desc", "szín.desc", "szin desc", "szín desc", "szin leiras", "szín leírás", "szin megnevezes", "szín megnevezés"},
        "exclude": {"leltarbol_ki", "leltárból_ki", "leltárból ki", "leltarbol ki"},
    }
    normalized_headers = [_normalize_header(value) for value in header_row]
    result: dict[str, int] = {}
    for key, names in aliases.items():
        normalized_names = {_normalize_header(name) for name in names}
        for index, header in enumerate(normalized_headers):
            if header in normalized_names:
                result[key] = index
                break
    return result


def _apply_material_book_qty_fallback(header_map: dict[str, int], header_row: tuple | list) -> None:
    """Anyagraktár files use the 3rd column for the bookkeeping quantity in practice.

    Keep explicit header detection first, but if the column name is not recognized,
    fall back to column C as long as it is not already used by another required field.
    """
    if "book_qty" in header_map or len(header_row) < 3:
        return
    fallback_index = 2
    used_by_required = {
        key: index
        for key, index in header_map.items()
        if key in {"part_number", "description", "icg_code", "exclude"}
    }
    if fallback_index not in used_by_required.values():
        header_map["book_qty"] = fallback_index


def _ensure_required_headers(header_map: dict[str, int]) -> None:
    missing = []
    for key, label in (("part_number", "Alkatr.-szám"), ("description", "Alkatr.-leírás"), ("icg_code", "ICG kód")):
        if key not in header_map:
            missing.append(label)
    if missing:
        raise ValueError("Hiányzó kötelező oszlop: " + ", ".join(missing))


def _ensure_semifinished_headers(header_map: dict[str, int]) -> None:
    missing = []
    for key, label in (("part_number", "Alkatr.-szám"), ("description", "Alkatr.-leírás")):
        if key not in header_map:
            missing.append(label)
    if "color_desc" not in header_map and "color_code" not in header_map:
        missing.append("SZIN vagy SZIN.Desc")
    if missing:
        raise ValueError("Hiányzó kötelező oszlop: " + ", ".join(missing))


def _row_from_values(header_map: dict[str, int], values: tuple | list) -> dict:
    def get_value(key: str) -> object:
        index = header_map.get(key)
        if index is None or index >= len(values):
            return ""
        return values[index]

    return {
        "part_number": get_value("part_number"),
        "description": get_value("description"),
        "book_qty": get_value("book_qty"),
        "icg_code": get_value("icg_code"),
        "exclude": get_value("exclude"),
    }


def _semifinished_row_from_values(header_map: dict[str, int], values: tuple | list) -> dict:
    def get_value(key: str) -> object:
        index = header_map.get(key)
        if index is None or index >= len(values):
            return ""
        return values[index]

    return {
        "part_number": get_value("part_number"),
        "description": get_value("description"),
        "color_code": get_value("color_code"),
        "color_desc": get_value("color_desc"),
        "exclude": get_value("exclude"),
    }


def _normalize_header(value: object) -> str:
    text = _clean_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    replacements = (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ö", "o"), ("ő", "o"), ("ú", "u"), ("ü", "u"), ("ű", "u"))
    for source, target in replacements:
        text = text.replace(source, target)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _clean_number_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return _format_quantity(float(value))
    return _clean_text(value)


def _is_excluded(value: object) -> bool:
    text = _clean_text(value).lower()
    return text not in {"", "0", "nem", "false", "no"}


def _row_id(part_number: str, category: str, index: str) -> str:
    import hashlib

    payload = f"{part_number}|{category}|{index}"
    return hashlib.sha1(payload.encode("utf-8", errors="ignore")).hexdigest()[:16]


def _category_key(value: object) -> str:
    text = _clean_text(value) or "Kategória nélkül"
    folded = _normalize_header(text)
    return re.sub(r"[^a-z0-9]+", "-", folded).strip("-") or "kategoria-nelkul"


def _category_summaries(rows: list[dict]) -> list[dict]:
    categories: dict[str, dict] = {}
    for row in rows:
        label = _clean_text(row.get("icg_code")) or "Kategória nélkül"
        key = _category_key(label)
        item = categories.setdefault(key, {"key": key, "label": label, "count": 0, "complete": True})
        item["count"] += 1
        if not str(row.get("input_qty", "")).strip():
            item["complete"] = False
    ordered = sorted(categories.values(), key=lambda item: _category_sort_key(item["label"]))
    total_complete = bool(rows) and all(str(row.get("input_qty", "")).strip() for row in rows)
    return [{"key": "all", "label": "Összes", "count": len(rows), "complete": total_complete}, *ordered]


def _summary_by_category(rows: list[dict]) -> list[dict]:
    categories: dict[str, dict] = {}
    for row in rows:
        label = _clean_text(row.get("icg_code")) or "Kategória nélkül"
        item = categories.setdefault(label, {"icg_code": label, "row_count": 0, "counted_count": 0, "total_qty": 0})
        item["row_count"] += 1
        value = _parse_non_negative_number(row.get("counted_qty") or row.get("input_qty"))
        if value is not None:
            item["counted_count"] += 1
            item["total_qty"] += value
    return sorted(categories.values(), key=lambda item: _category_sort_key(item["icg_code"]))


def _row_sort_key(row: dict) -> tuple:
    return (_category_sort_key(row.get("icg_code")), _clean_text(row.get("part_number")).lower())


def _category_sort_key(value: object) -> tuple:
    text = _clean_text(value)
    return (_normalize_header(text), text)


def _find_row(session: dict, row_id: str) -> dict | None:
    clean_id = _clean_text(row_id)
    for row in session.get("rows", []):
        if isinstance(row, dict) and str(row.get("row_id", "")) == clean_id:
            return row
    return None


def _touch_session(session: dict) -> None:
    session["updated_at"] = datetime.now().isoformat(timespec="seconds")


def _parse_non_negative_number(value: object) -> float | None:
    text = _clean_text(value)
    if not text:
        return None
    text = text.replace(" ", "").replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    if number < 0:
        return None
    return number


def _format_quantity(value: float | int) -> str:
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.3f}".rstrip("0").rstrip(".")


def _excel_quantity(row: dict) -> int | float:
    value = _parse_non_negative_number(row.get("counted_qty") or row.get("input_qty"))
    if value is None:
        return 0
    return int(value) if float(value).is_integer() else value


def _style_sheet(sheet, widths: tuple[int, ...]) -> None:
    if Font is None or PatternFill is None or Alignment is None:
        return
    header_fill = PatternFill("solid", fgColor="E8F7F1")
    header_font = Font(bold=True, color="0F172A")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
