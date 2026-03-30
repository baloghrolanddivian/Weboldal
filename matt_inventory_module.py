from __future__ import annotations

import csv
import io
import json
import re
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:  # pragma: no cover - optional dependency handling
    Workbook = None
    load_workbook = None


MATT_INVENTORY_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}
FAMILY_STOPWORDS = {
    "folias",
    "foliasfr",
    "fr",
    "egyedi",
    "rossz",
    "regi",
    "tak",
    "taks",
    "sarok",
    "blende",
    "also",
    "felso",
    "ives",
    "ajto",
    "marassal",
    "mellved",
    "matt",
    "supermatt",
    "rusztikus",
    "csak",
    "elmarassal",
    "nf",
    "felso",
}
ODS_TABLE_NS = {
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}


@dataclass(slots=True)
class MattInventoryGroup:
    family: str
    color: str
    quantity: Decimal
    total_value: Decimal
    line_count: int
    part_count: int

    @property
    def label(self) -> str:
        return f"{self.family} {self.color}".strip()

    @property
    def average_cost(self) -> Decimal:
        if self.quantity <= 0:
            return Decimal("0")
        return (self.total_value / self.quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def to_dict(self) -> dict:
        return {
            "family": self.family,
            "color": self.color,
            "quantity": _decimal_to_string(self.quantity),
            "total_value": _decimal_to_string(self.total_value),
            "line_count": self.line_count,
            "part_count": self.part_count,
        }

    @classmethod
    def from_dict(cls, payload: dict) -> "MattInventoryGroup":
        return cls(
            family=str(payload.get("family", "")).strip(),
            color=str(payload.get("color", "")).strip(),
            quantity=_decimal_from_value(payload.get("quantity")),
            total_value=_decimal_from_value(payload.get("total_value")),
            line_count=max(0, int(payload.get("line_count", 0) or 0)),
            part_count=max(0, int(payload.get("part_count", 0) or 0)),
        )


@dataclass(slots=True)
class MattPriceInfo:
    cost: Decimal
    storage_limit: Decimal
    safety_stock: Decimal


@dataclass(slots=True)
class MattInventoryReport:
    generated_at: str
    price_source_name: str
    stock_source_name: str
    groups: list[MattInventoryGroup]
    total_quantity: Decimal
    total_value: Decimal
    matched_row_count: int
    missing_codes: list[str]
    safety_exceeded_count: int
    storage_exceeded_count: int

    def to_dict(self) -> dict:
        return {
            "generated_at": self.generated_at,
            "price_source_name": self.price_source_name,
            "stock_source_name": self.stock_source_name,
            "groups": [group.to_dict() for group in self.groups],
            "total_quantity": _decimal_to_string(self.total_quantity),
            "total_value": _decimal_to_string(self.total_value),
            "matched_row_count": self.matched_row_count,
            "missing_codes": list(self.missing_codes),
            "safety_exceeded_count": self.safety_exceeded_count,
            "storage_exceeded_count": self.storage_exceeded_count,
        }

    @classmethod
    def from_dict(cls, payload: dict) -> "MattInventoryReport":
        groups = [MattInventoryGroup.from_dict(item) for item in payload.get("groups", []) if isinstance(item, dict)]
        return cls(
            generated_at=str(payload.get("generated_at", "")).strip(),
            price_source_name=str(payload.get("price_source_name", "")).strip(),
            stock_source_name=str(payload.get("stock_source_name", "")).strip(),
            groups=groups,
            total_quantity=_decimal_from_value(payload.get("total_quantity")),
            total_value=_decimal_from_value(payload.get("total_value")),
            matched_row_count=max(0, int(payload.get("matched_row_count", 0) or 0)),
            missing_codes=[str(item).strip() for item in payload.get("missing_codes", []) if str(item).strip()],
            safety_exceeded_count=max(0, int(payload.get("safety_exceeded_count", 0) or 0)),
            storage_exceeded_count=max(0, int(payload.get("storage_exceeded_count", 0) or 0)),
        )


def load_report_from_path(path: Path) -> MattInventoryReport | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    return MattInventoryReport.from_dict(payload)


def save_report_to_path(path: Path, report: MattInventoryReport) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")


def build_matt_inventory_report(
    *,
    price_name: str,
    price_bytes: bytes,
    stock_name: str,
    stock_bytes: bytes,
) -> MattInventoryReport:
    price_map = _read_price_map(price_name, price_bytes)
    stock_rows = _read_stock_rows(stock_name, stock_bytes)
    family_code_map = _derive_family_code_map(stock_rows)

    grouped: dict[tuple[str, str], dict] = {}
    total_quantity = Decimal("0")
    total_value = Decimal("0")
    matched_row_count = 0
    missing_codes: list[str] = []
    seen_missing: set[str] = set()
    safety_exceeded_count = 0
    storage_exceeded_count = 0

    for row in stock_rows:
        if row["quantity"] <= 0:
            continue

        part_number = row["part_number"]
        price_info = _lookup_price(part_number, price_map)
        if price_info is None:
            if part_number not in seen_missing:
                seen_missing.add(part_number)
                missing_codes.append(part_number)
            continue
        unit_cost = price_info.cost

        family = _family_display_name(row["family_code"], row["description"], row["color"], family_code_map)
        color = _normalize_color_name(row["color"])
        key = (family, color)
        bucket = grouped.setdefault(
            key,
            {
                "family": family,
                "color": color,
                "quantity": Decimal("0"),
                "total_value": Decimal("0"),
                "line_count": 0,
                "parts": set(),
            },
        )
        bucket["quantity"] += row["quantity"]
        bucket["total_value"] += row["quantity"] * unit_cost
        bucket["line_count"] += 1
        bucket["parts"].add(part_number)
        total_quantity += row["quantity"]
        total_value += row["quantity"] * unit_cost
        matched_row_count += 1
        threshold_eligible = price_info.safety_stock > 1
        if threshold_eligible and row["quantity"] > price_info.safety_stock:
            safety_exceeded_count += 1
        if threshold_eligible and row["quantity"] > price_info.storage_limit:
            storage_exceeded_count += 1

    groups = [
        MattInventoryGroup(
            family=str(item["family"]).strip(),
            color=str(item["color"]).strip(),
            quantity=_money_round(item["quantity"]),
            total_value=_money_round(item["total_value"]),
            line_count=int(item["line_count"]),
            part_count=len(item["parts"]),
        )
        for item in grouped.values()
    ]
    groups.sort(key=lambda group: (_fold_text(group.family), _fold_text(group.color)))

    return MattInventoryReport(
        generated_at=datetime.now().isoformat(timespec="seconds"),
        price_source_name=Path(price_name).name.strip(),
        stock_source_name=Path(stock_name).name.strip(),
        groups=groups,
        total_quantity=_money_round(total_quantity),
        total_value=_money_round(total_value),
        matched_row_count=matched_row_count,
        missing_codes=missing_codes,
        safety_exceeded_count=safety_exceeded_count,
        storage_exceeded_count=storage_exceeded_count,
    )


def build_matt_inventory_alert_workbook(
    *,
    price_name: str,
    price_bytes: bytes,
    stock_name: str,
    stock_bytes: bytes,
) -> bytes:
    if Workbook is None:
        raise RuntimeError("Az Excel riport készítéséhez hiányzik az openpyxl csomag.")

    price_map = _read_price_map(price_name, price_bytes)
    stock_rows = _read_stock_rows(stock_name, stock_bytes)
    family_code_map = _derive_family_code_map(stock_rows)
    element_map = _load_element_map()

    safety_rows: list[dict] = []
    storage_rows: list[dict] = []

    for row in stock_rows:
        if row["quantity"] <= 0:
            continue
        price_info = _lookup_price(row["part_number"], price_map)
        if price_info is None:
            continue

        family = _family_display_name(row["family_code"], row["description"], row["color"], family_code_map)
        color = _normalize_color_name(row["color"])
        front_size = _extract_front_size(row["description"], row["part_number"])
        related_elements = element_map.get(front_size, [])
        value = _money_round(row["quantity"] * price_info.cost)
        base = {
            "part_number": row["part_number"],
            "description": row["description"],
            "family": family,
            "color": color,
            "front_size": front_size,
            "elements": ", ".join(related_elements),
            "quantity": row["quantity"],
            "material_cost": _money_round(price_info.cost),
            "stock_value": value,
        }

        if price_info.safety_stock > 1 and row["quantity"] > price_info.safety_stock:
            safety_rows.append(
                {
                    **base,
                    "threshold": _money_round(price_info.safety_stock),
                    "excess": _money_round(row["quantity"] - price_info.safety_stock),
                }
            )
        if price_info.safety_stock > 1 and row["quantity"] > price_info.storage_limit:
            storage_rows.append(
                {
                    **base,
                    "threshold": _money_round(price_info.storage_limit),
                    "excess": _money_round(row["quantity"] - price_info.storage_limit),
                }
            )

    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Biztonsági felett"
    _populate_alert_sheet(first_sheet, safety_rows, "Biztonsági készlet felett")
    second_sheet = workbook.create_sheet("Tárolható felett")
    _populate_alert_sheet(second_sheet, storage_rows, "Tárolható mennyiség felett")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_runtime_upload(path: Path, file_name: str, payload: bytes) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = Path(file_name).suffix.lower() or ".xlsx"
    stored_path = path.with_suffix(suffix)
    stored_path.write_bytes(payload)
    return stored_path


def file_name_allowed(file_name: str) -> bool:
    return Path(file_name).suffix.lower() in MATT_INVENTORY_ALLOWED_EXTENSIONS


def read_bytes_if_exists(path: Path) -> bytes | None:
    if not path.exists():
        return None
    return path.read_bytes()


def _read_price_map(file_name: str, payload: bytes) -> dict[str, MattPriceInfo]:
    rows = _read_rows(file_name, payload)
    if not rows:
        raise ValueError("A fix ártábla üres.")

    header = rows[0]
    header_map = _build_header_map(header)
    part_index = _find_header_index(header_map, ("alkatr", "szam"))
    storage_index = _find_header_index(header_map, ("tarolh", "menny"))
    safety_index = _find_header_index(header_map, ("biztonsagi", "keszlet"))
    cost_index = _find_header_index(header_map, ("anyag", "koltseg"))
    if part_index is None or cost_index is None or storage_index is None or safety_index is None:
        raise ValueError("A fix ártáblában az 'Alkatr.-szám', 'Tárolh.menny.', 'Biztonsági készlet' és 'Anyag-Költség' oszlopok szükségesek.")

    result: dict[str, MattPriceInfo] = {}
    for row in rows[1:]:
        if max(part_index, storage_index, safety_index, cost_index) >= len(row):
            continue
        part_number = _normalize_part_number(row[part_index])
        if not part_number:
            continue
        cost = _decimal_from_value(row[cost_index])
        if cost <= 0:
            continue
        result[part_number] = MattPriceInfo(
            cost=cost,
            storage_limit=_decimal_from_value(row[storage_index]),
            safety_stock=_decimal_from_value(row[safety_index]),
        )

    if not result:
        raise ValueError("A fix ártáblában nem találtam használható alkatrészszám + küszöb + anyagköltség sorokat.")
    return result


def _read_stock_rows(file_name: str, payload: bytes) -> list[dict]:
    rows = _read_rows(file_name, payload)
    if not rows:
        raise ValueError("A napi készletfájl üres.")

    header = rows[0]
    header_map = _build_header_map(header)
    part_index = _find_header_index(header_map, ("alkatr", "szam"))
    desc_index = _find_header_index(header_map, ("alkatr", "leiras"))
    qty_index = _find_header_index(header_map, ("rend", "all"))
    color_index = _find_header_index(header_map, ("szin", "desc"))

    if None in {part_index, desc_index, qty_index, color_index}:
        raise ValueError("A napi készletfájlban az első 4 oszlopnak ezeknek kell lennie: alkatrész szám, leírás, mennyiség, szín.")

    items: list[dict] = []
    for row in rows[1:]:
        if max(part_index, desc_index, qty_index, color_index) >= len(row):
            continue
        part_number = _normalize_part_number(row[part_index])
        if not part_number:
            continue
        quantity = _decimal_from_value(row[qty_index])
        description = str(row[desc_index] or "").strip()
        color = str(row[color_index] or "").strip()
        items.append(
            {
                "part_number": part_number,
                "description": description,
                "quantity": quantity,
                "color": color,
                "family_code": _family_code_from_part(part_number),
            }
        )

    if not items:
        raise ValueError("A napi készletfájlban nem találtam használható sorokat.")
    return items


def _read_rows(file_name: str, payload: bytes) -> list[list[object]]:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise RuntimeError("Az Excel feldolgozáshoz hiányzik az openpyxl csomag.")
        workbook = load_workbook(io.BytesIO(payload), data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    if suffix == ".csv":
        text = _decode_csv_bytes(payload)
        return [list(row) for row in csv.reader(io.StringIO(text))]

    raise ValueError("Csak XLSX, XLSM vagy CSV fájlokat tudok feldolgozni.")


def _decode_csv_bytes(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="ignore")


def _build_header_map(header_row: list[object]) -> dict[int, str]:
    return {index: _normalize_header(value) for index, value in enumerate(header_row)}


def _find_header_index(header_map: dict[int, str], terms: tuple[str, ...]) -> int | None:
    for index, normalized in header_map.items():
        if all(term in normalized for term in terms):
            return index
    return None


def _normalize_header(value: object) -> str:
    text = _fold_text(str(value or ""))
    return re.sub(r"[^a-z0-9]+", "", text)


def _normalize_part_number(value: object) -> str:
    return str(value or "").strip().upper()


def _decimal_from_value(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "")
    if not text:
        return Decimal("0")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _money_round(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _decimal_to_string(value: Decimal) -> str:
    normalized = _money_round(value)
    text = format(normalized, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _family_code_from_part(part_number: str) -> str:
    parts = [part for part in part_number.split("_") if part]
    return parts[1] if len(parts) > 1 else ""


def _lookup_price(part_number: str, price_map: dict[str, MattPriceInfo]) -> MattPriceInfo | None:
    for alias in _price_aliases(part_number):
        unit_cost = price_map.get(alias)
        if unit_cost is not None and unit_cost.cost > 0:
            return unit_cost
    return None


def _price_aliases(part_number: str) -> list[str]:
    clean_part = _normalize_part_number(part_number)
    candidates = [clean_part]
    if clean_part.startswith("NFAU_"):
        candidates.append(clean_part.replace("NFAU_", "NFA_", 1))
        candidates.append(clean_part.replace("NFAU_", "NFAY_", 1))
    if clean_part.startswith("NFAY_"):
        candidates.append(clean_part.replace("NFAY_", "NFA_", 1))

    reordered = re.sub(r"^(.*)_KA_([^_]+_[^_]+)_NFA$", r"\1_NFA_KA_\2", clean_part)
    if reordered != clean_part:
        candidates.append(reordered)

    unique_candidates: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        if candidate and candidate not in seen:
            seen.add(candidate)
            unique_candidates.append(candidate)
    return unique_candidates


def _derive_family_code_map(stock_rows: list[dict]) -> dict[str, str]:
    candidates: dict[str, Counter] = defaultdict(Counter)
    for row in stock_rows:
        family_code = str(row.get("family_code", "")).strip()
        candidate = _family_candidate_from_description(str(row.get("description", "")), str(row.get("color", "")))
        if family_code and candidate:
            candidates[family_code][candidate] += 1

    return {
        family_code: counter.most_common(1)[0][0]
        for family_code, counter in candidates.items()
        if counter
    }


def _family_display_name(family_code: str, description: str, color: str, code_map: dict[str, str]) -> str:
    direct = code_map.get(family_code)
    if direct:
        return direct

    candidate = _family_candidate_from_description(description, color)
    if candidate:
        return candidate

    if family_code:
        return family_code.title()
    return "Ismeretlen"


def _family_candidate_from_description(description: str, color: str) -> str:
    working = str(description or "").strip()
    if not working:
        return ""

    working = re.sub(r"\s+\d+x\d+x\d+(?:\s+[A-Z])?\s*$", "", working, flags=re.IGNORECASE)
    if color:
        position = working.casefold().find(str(color).strip().casefold())
        if position != -1:
            working = working[:position].strip()

    direct_patterns = (
        r"fr\.?\s*([A-ZÁÉÍÓÖŐÚÜŰ][a-záéíóöőúüű]+)",
        r"Tak\.?s?\s*([A-ZÁÉÍÓÖŐÚÜŰ][a-záéíóöőúüű]+)",
        r"f\.\s*([A-ZÁÉÍÓÖŐÚÜŰ][a-záéíóöőúüű]+)",
    )
    for pattern in direct_patterns:
        match = re.search(pattern, working, flags=re.IGNORECASE)
        if match:
            return _clean_family_name(match.group(1))

    candidates = re.findall(r"[A-ZÁÉÍÓÖŐÚÜŰ][a-záéíóöőúüű]+", working)
    for candidate in reversed(candidates):
        if _fold_text(candidate) not in FAMILY_STOPWORDS:
            return _clean_family_name(candidate)
    return ""


def _clean_family_name(value: str) -> str:
    clean_value = re.sub(r"\s+", " ", str(value or "")).strip(" .-_")
    if not clean_value:
        return ""
    return clean_value[0].upper() + clean_value[1:]


def _normalize_color_name(value: str) -> str:
    color = str(value or "").strip()
    if not color or _fold_text(color) == "nincs":
        return "Ismeretlen szín"

    color = re.sub(r"^Mf\.\s*", "Matt ", color, flags=re.IGNORECASE)
    color = re.sub(r"^SM\.\s*", "Supermatt ", color, flags=re.IGNORECASE)
    color = re.sub(r"\bfóliás\b", "", color, flags=re.IGNORECASE)
    color = re.sub(r"\bkr\.\b", "", color, flags=re.IGNORECASE)
    color = re.sub(r"capuccino", "cappuccino", color, flags=re.IGNORECASE)
    color = re.sub(r"\s+", " ", color).strip(" .")
    if not color:
        return "Ismeretlen szín"
    return " ".join(part.capitalize() for part in color.split())


def _populate_alert_sheet(sheet, rows: list[dict], title: str) -> None:
    title_fill = PatternFill(fill_type="solid", fgColor="0F172A")
    title_font = Font(color="FFFFFF", bold=True, size=12)
    header_fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    header_font = Font(color="0F172A", bold=True)

    headers = [
        "Alkatrész-szám",
        "Leírás",
        "Modell",
        "Szín",
        "Front méret",
        "Elemek",
        "Készlet",
        "Küszöb",
        "Eltérés",
        "Anyagköltség",
        "Készletérték",
    ]

    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(row=1, column=1)
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    sheet.append(headers)
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=index)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for entry in rows:
        sheet.append(
            [
                entry["part_number"],
                entry["description"],
                entry["family"],
                entry["color"],
                entry["front_size"],
                entry["elements"],
                float(entry["quantity"]),
                float(entry["threshold"]),
                float(entry["excess"]),
                float(entry["material_cost"]),
                float(entry["stock_value"]),
            ]
        )

    widths = {
        1: 30,
        2: 44,
        3: 16,
        4: 22,
        5: 14,
        6: 30,
        7: 12,
        8: 12,
        9: 12,
        10: 14,
        11: 16,
    }
    for column_index, width in widths.items():
        sheet.column_dimensions[_excel_column_name(column_index)].width = width

    for row in sheet.iter_rows(min_row=3, min_col=7, max_col=11):
        for cell in row:
            cell.number_format = '#,##0.00'

    sheet.freeze_panes = "A3"


def _load_element_map() -> dict[str, list[str]]:
    source_path = _latest_element_map_source()
    if source_path is None:
        return {}
    try:
        return _read_element_map_from_ods(source_path)
    except Exception:
        return {}


def _latest_element_map_source() -> Path | None:
    downloads_dir = Path.home() / "Downloads"
    candidates: list[Path] = []
    for pattern in ("Elem fogyás 2025*.ods", "Elem fogyas 2025*.ods"):
        candidates.extend(downloads_dir.glob(pattern))
    if not candidates:
        return None
    return max(candidates, key=lambda item: item.stat().st_mtime_ns)


def _read_element_map_from_ods(path: Path) -> dict[str, list[str]]:
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("content.xml"))

    target_table = None
    for table in root.findall(".//table:table", ODS_TABLE_NS):
        table_name = str(table.attrib.get(f"{{{ODS_TABLE_NS['table']}}}name", "")).strip()
        if _fold_text(table_name) == "meret":
            target_table = table
            break
    if target_table is None:
        return {}

    result: dict[str, list[str]] = {}
    for row in target_table.findall("table:table-row", ODS_TABLE_NS):
        values = _ods_row_values(row)
        if not any(value.strip() for value in values):
            continue
        first_cell = str(values[0] or "").strip()
        if _fold_text(first_cell) in {"", "meret"}:
            continue
        size_key = _normalize_front_size(first_cell)
        if not size_key:
            continue
        elements = [str(value).strip() for value in values[2:] if str(value).strip()]
        if elements:
            result[size_key] = elements
    return result


def _ods_row_values(row) -> list[str]:
    values: list[str] = []
    repeat_attr = f"{{{ODS_TABLE_NS['table']}}}number-columns-repeated"
    for cell in row.findall("table:table-cell", ODS_TABLE_NS):
        repeat_count = int(cell.attrib.get(repeat_attr, "1"))
        texts: list[str] = []
        for paragraph in cell.findall("text:p", ODS_TABLE_NS):
            text_value = "".join(paragraph.itertext()).strip()
            if text_value:
                texts.append(text_value)
        cell_value = " ".join(texts).strip()
        values.extend([cell_value] * repeat_count)
    return values


def _extract_front_size(description: str, part_number: str) -> str:
    for source in (str(description or ""), str(part_number or "")):
        match = re.search(r"(\d{2,4}\s*x\s*\d{2,4})\s*x\s*\d{1,2}\b", source, flags=re.IGNORECASE)
        if match:
            return _normalize_front_size(match.group(1))
        match = re.search(r"(\d{2,4}\s*x\s*\d{2,4})\b", source, flags=re.IGNORECASE)
        if match:
            return _normalize_front_size(match.group(1))
    return ""


def _normalize_front_size(value: str) -> str:
    clean_value = re.sub(r"\s+", "", str(value or "")).lower()
    match = re.match(r"^(\d{2,4})x(\d{2,4})$", clean_value)
    if not match:
        return ""
    return f"{match.group(1)}x{match.group(2)}"


def _excel_column_name(index: int) -> str:
    result = ""
    current = int(index)
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _fold_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(char for char in normalized if not unicodedata.combining(char)).casefold().strip()
