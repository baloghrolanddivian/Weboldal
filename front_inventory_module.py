from __future__ import annotations

import csv
import io
import json
import re
import secrets
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None


FRONT_INVENTORY_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}
SERIAL_SIZES_DATA_PATH = Path(__file__).resolve().parent / "data" / "front-inventory-serial-sizes.json"
NETTFRONT_TRANSLATIONS_DATA_PATH = Path(__file__).resolve().parent / "data" / "nettfront-translations.json"
FRONT_INVENTORY_INSIGHT_TEMPLATE_OVERRIDE_PATH = Path(__file__).resolve().parent / "runtime" / "front-leltar" / "insight-minta.xlsx"
FRONT_INVENTORY_INSIGHT_TEMPLATE_PATH = Path(__file__).resolve().parent / "data" / "nettfront-alkatreszek.xlsx"


def file_name_allowed(file_name: str) -> bool:
    return Path(file_name or "").suffix.lower() in FRONT_INVENTORY_ALLOWED_EXTENSIONS


def read_bytes_if_exists(path: Path) -> bytes | None:
    if not path.exists():
        return None
    return path.read_bytes()


def write_runtime_upload(base_path: Path, file_name: str, payload: bytes) -> Path:
    suffix = Path(file_name or "").suffix.lower() or ".bin"
    target_path = base_path.with_suffix(suffix)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(payload)
    return target_path


def load_session_from_path(path: Path) -> dict | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    if _repair_session(payload, path):
        save_session_to_path(path, payload)
    return payload


def save_session_to_path(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_front_inventory_session(file_name: str, payload: bytes) -> dict:
    serial_sizes = _load_serial_sizes()
    stock_rows = _read_stock_rows(file_name, payload, serial_sizes)

    rows: list[dict] = []
    for item in stock_rows:
        rows.append(_build_inventory_session_row(item, serial_sizes))

    rows.sort(key=_row_sort_key)

    return {
        "session_id": secrets.token_hex(6),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "source_name": Path(file_name).name,
        "phase": 0,
        "phase_label": "Számlálás",
        "finalized_at": "",
        "serial_sizes": sorted(serial_sizes, key=_size_sort_key),
        "rows": rows,
    }


def update_row_input(session: dict, row_id: str, raw_value: str) -> tuple[bool, str]:
    row = _find_row(session, row_id)
    if row is None:
        return False, "A kiválasztott frontsort nem találom."
    if str(session.get("phase")) == "finalized":
        return False, "A leltár már le van zárva."
    if row.get("status") != "pending":
        return False, "Ez a sor már lezárt állapotban van."

    clean_value = str(raw_value or "").strip()
    if not clean_value:
        row["input_qty"] = ""
        _touch_session(session)
        return True, ""

    parsed_value = _parse_non_negative_int(clean_value)
    if parsed_value is None:
        return False, "Csak nem negatív egész darabszám adható meg."

    row["input_qty"] = str(parsed_value)
    _touch_session(session)
    return True, ""


def summarize_missing_inputs(session: dict) -> dict:
    if str(session.get("phase")) == "finalized":
        return {"total_missing": 0, "categories": [], "rows": []}

    phase_value = session.get("phase", 0)
    phase = 2 if str(phase_value) == "finalized" else int(phase_value or 0)
    active_rows = _active_rows_for_phase(session, phase)
    missing_rows = [row for row in active_rows if _row_input_value(row) is None]

    category_counts: dict[str, int] = {}
    for row in missing_rows:
        category_key = str(row.get("category") if row.get("is_serial") else "egyedi")
        category_counts[category_key] = category_counts.get(category_key, 0) + 1

    categories = [
        {"key": key, "count": count}
        for key, count in sorted(category_counts.items(), key=lambda item: _inventory_category_sort_key(item[0]))
    ]
    rows = [
        {
            "part_number": str(row.get("part_number", "")),
            "description": str(row.get("description", "")),
            "color": str(row.get("color", "")),
            "category": str(row.get("category") if row.get("is_serial") else "egyedi"),
        }
        for row in missing_rows[:40]
    ]
    return {
        "total_missing": len(missing_rows),
        "categories": categories,
        "rows": rows,
    }


def _set_worker_alert(session: dict, title: str, message: str) -> None:
    session["worker_alert"] = {
        "id": datetime.now().isoformat(timespec="seconds"),
        "title": str(title or "").strip(),
        "message": str(message or "").strip(),
    }


def build_inventory_check_workbook(session: dict, mode: str = "check", treat_missing_as_zero: bool = False) -> tuple[bytes | None, str, int]:
    if Workbook is None:
        return None, "", 0

    phase_value = session.get("phase", 0)
    phase = 2 if str(phase_value) == "finalized" else int(phase_value or 0)
    all_rows = sorted([row for row in session.get("rows", []) if _is_visible_inventory_row(row)], key=_row_sort_key)
    if mode == "finalize":
        report_label = "veglegesites"
    elif phase == 0:
        report_label = "elso-ellenorzes"
    elif phase == 1:
        report_label = "masodik-ellenorzes"
    else:
        report_label = "ellenorzes"

    workbook = Workbook()
    report_rows = [_build_export_row(row, phase=phase, mode=mode, treat_missing_as_zero=treat_missing_as_zero) for row in all_rows]

    if mode == "finalize":
        _fill_finalize_workbook(workbook, report_rows)
    else:
        sheet = workbook.active
        sheet.title = "Ellenorzes"
        if phase == 1:
            headers = ["Alkatresz szam", "Leiras", "Elso szamlalas", "Masodik szamlalas", "Raktari darabszam"]
        else:
            headers = ["Alkatresz szam", "Leiras", "Beirt darabszam", "Raktari darabszam"]
        sheet.append(headers)
        if report_rows:
            for item in report_rows:
                if phase == 1:
                    sheet.append(
                        [
                            item["part_number"],
                            item["description"],
                            _excel_cell_int(item.get("first_check_qty")),
                            _excel_cell_int(item.get("second_check_qty")),
                            item["stock_qty"],
                        ]
                    )
                else:
                    sheet.append(
                        [
                            item["part_number"],
                            item["description"],
                            _excel_cell_int(item.get("current_count_qty")),
                            item["stock_qty"],
                        ]
                    )
        else:
            sheet.append(["", "Nincs elteres ebben az ellenorzesi korben."] + [""] * (len(headers) - 2))

        for width, column in zip((26, 56, 18, 18, 18), ("A", "B", "C", "D", "E")):
            sheet.column_dimensions[column].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), f"front-leltar-{report_label}.xlsx", len(all_rows)


def build_front_inventory_insight_artifacts(session: dict) -> dict[str, object]:
    if load_workbook is None:
        raise RuntimeError("Az inSight Excel generalasahoz hianyzik az openpyxl csomag.")

    template_path = _resolve_front_inventory_insight_template_path()
    if template_path is None:
        raise FileNotFoundError("Nem talaltam az inSight minta Excelt.")

    workbook = load_workbook(template_path)
    sheet = workbook.active
    part_column_index, qty_column_index = _front_inventory_insight_columns(sheet)
    final_counts = _front_inventory_final_count_map(session)
    ordered_quantities: list[int] = []
    ordered_parts: list[str] = []
    seen_parts: set[str] = set()

    qty_header = sheet.cell(row=1, column=qty_column_index + 1).value
    if not str(qty_header or "").strip():
        sheet.cell(row=1, column=qty_column_index + 1).value = "Szamolt mennyiseg"

    for row_index in range(2, sheet.max_row + 1):
        part_value = sheet.cell(row=row_index, column=part_column_index + 1).value
        part_number = _normalize_part_number(part_value)
        if not part_number:
            continue
        quantity = int(final_counts.get(part_number, 0) or 0)
        sheet.cell(row=row_index, column=qty_column_index + 1).value = quantity
        ordered_parts.append(part_number)
        ordered_quantities.append(quantity)
        seen_parts.add(part_number)

    missing_parts = sorted(part_number for part_number in final_counts if part_number not in seen_parts)
    if missing_parts:
        missing_sheet = workbook.create_sheet("Nincs a mintaban")
        missing_sheet.append(["Alkatresz szam", "Darabszam"])
        for part_number in missing_parts:
            missing_sheet.append([part_number, int(final_counts.get(part_number, 0) or 0)])

    buffer = io.BytesIO()
    workbook.save(buffer)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return {
        "workbook": buffer.getvalue(),
        "workbook_name": f"front-leltar-insight-bevetelezes-{stamp}.xlsx",
        "script": _build_front_inventory_insight_ahk_script(ordered_quantities).encode("utf-8"),
        "script_name": f"front-leltar-insight-bevetelezes-{stamp}.ahk",
        "row_count": len(ordered_parts),
        "matched_count": len(final_counts) - len(missing_parts),
        "missing_parts": missing_parts,
        "template_name": template_path.name,
    }


def run_inventory_check(session: dict, allow_missing: bool = False) -> tuple[bool, str]:
    if str(session.get("phase")) == "finalized":
        return False, "A leltár már le van zárva."

    phase = int(session.get("phase", 0) or 0)
    active_rows = _active_rows_for_phase(session, phase)
    if not active_rows:
        session["phase"] = 2
        session["phase_label"] = "Veglegesites"
        _touch_session(session)
        return True, "Nincs tovabbi ellenorizendo front, a leltar veglegesitheto."

    missing_rows = [row for row in active_rows if _row_input_value(row) is None]
    missing_count = len(missing_rows)
    if missing_count and not allow_missing:
        return False, f"Meg {missing_count} frontnal nincs beirva a darabszam."

    if phase == 0:
        mismatch_count = 0
        for row in active_rows:
            entered_qty = _row_input_value(row)
            if entered_qty is None:
                if allow_missing:
                    entered_qty = 0
                assert entered_qty is not None
            row["first_check_qty"] = entered_qty
            if entered_qty == int(row.get("stock_qty", 0) or 0):
                row["resolved_qty"] = entered_qty
                row["status"] = "resolved"
            else:
                row["review_level"] = 1
                mismatch_count += 1
        _clear_all_pending_inputs(session)
        session["phase"] = 1 if mismatch_count else 2
        session["phase_label"] = "Elso ellenorzes" if mismatch_count else "Veglegesites"
        _touch_session(session)
        if mismatch_count:
            _set_worker_alert(session, "Ellenorzes kesz", f"{mismatch_count} tetelnel elteres talalhato, ujraszamolas szukseges.")
            if missing_count and allow_missing:
                return True, f"{mismatch_count} front eltert a keszlettol, {missing_count} kitoltetlen sor pedig atkerult a kovetkezo korbe."
            return True, f"{mismatch_count} front eltert a keszlettol, ezeket ujra kell szamolni."
        _set_worker_alert(session, "Ellenorzes kesz", "Nem talaltam elterest, a leltar veglegesitheto.")
        if missing_count and allow_missing:
            return True, f"Nincs elteres, de {missing_count} kitoltetlen sor atkerult a kovetkezo korbe."
        return True, "Minden front egyezik a keszlettel, a leltar veglegesitheto."

    high_diff_count = 0
    for row in active_rows:
        entered_qty = _row_input_value(row)
        if entered_qty is None:
            if allow_missing:
                entered_qty = 0
            assert entered_qty is not None
        row["second_check_qty"] = entered_qty
        stock_qty = int(row.get("stock_qty", 0) or 0)
        if abs(entered_qty - stock_qty) > 5:
            row["review_level"] = 2
            high_diff_count += 1
        else:
            row["resolved_qty"] = entered_qty
            row["status"] = "resolved"

    _clear_all_pending_inputs(session)
    session["phase"] = 2
    session["phase_label"] = "Veglegesites"
    _touch_session(session)
    if high_diff_count:
        _set_worker_alert(session, "Ellenorzes kesz", f"{high_diff_count} tetelnel tovabbra is nagy elteres van, ujraszamolas szukseges.")
        if missing_count and allow_missing:
            return True, f"{high_diff_count} frontnal 5 darabnal nagyobb elteres maradt, {missing_count} kitoltetlen sor pedig tovabbra is nyitott."
        return True, f"{high_diff_count} frontnal 5 darabnal nagyobb elteres maradt, ezeket meg egyszer ellenorizni kell."
    _set_worker_alert(session, "Ellenorzes kesz", "Az ujraellenorzes befejezodott, a leltar veglegesitheto.")
    if missing_count and allow_missing:
        return True, f"Az eltero frontok ujraellenorzese kesz, de {missing_count} kitoltetlen sor tovabbra is nyitott."
    return True, "Az eltero frontok ujraellenorzese kesz, a leltar veglegesitheto."


def finalize_inventory(session: dict, allow_missing: bool = False) -> tuple[bool, str]:
    if str(session.get("phase")) == "finalized":
        return False, "A leltar mar le van zarva."

    active_rows = _active_rows_for_phase(session, 2)
    missing_count = sum(1 for row in active_rows if _row_input_value(row) is None)
    if missing_count and not allow_missing:
        return False, f"Meg {missing_count} frontnal nincs vegleges darabszam."

    for row in active_rows:
        entered_qty = _row_input_value(row)
        if entered_qty is None:
            if allow_missing:
                entered_qty = 0
            assert entered_qty is not None
        row["final_check_qty"] = entered_qty
        row["resolved_qty"] = entered_qty
        row["status"] = "resolved"

    _clear_all_pending_inputs(session)
    session["phase"] = "finalized"
    session["phase_label"] = "Lezarva"
    session["finalized_at"] = datetime.now().isoformat(timespec="seconds")
    _touch_session(session)
    return True, "A leltar lezarult."


def _clear_all_pending_inputs(session: dict) -> None:
    for row in session.get("rows", []):
        row["input_qty"] = ""


def _build_export_row(row: dict, phase: int, mode: str, treat_missing_as_zero: bool) -> dict:
    current_input = _row_input_value(row)
    if current_input is None and treat_missing_as_zero and str(row.get("status", "")) != "resolved":
        current_input = 0

    export_row = {
        "part_number": str(row.get("part_number", "")).strip(),
        "description": str(row.get("description", "")).strip(),
        "stock_qty": int(row.get("stock_qty", 0) or 0),
        "first_check_qty": _parse_non_negative_int(row.get("first_check_qty")),
        "second_check_qty": _parse_non_negative_int(row.get("second_check_qty")),
        "final_check_qty": _parse_non_negative_int(row.get("final_check_qty")),
        "resolved_qty": _parse_non_negative_int(row.get("resolved_qty")),
        "current_count_qty": None,
    }

    if mode == "finalize":
        export_row["final_count_qty"] = _final_effective_qty(row)
        return export_row

    if phase == 0:
        export_row["current_count_qty"] = current_input
    elif phase == 1 and int(row.get("review_level", 0) or 0) == 1 and str(row.get("status", "")) != "resolved":
        export_row["second_check_qty"] = current_input
    else:
        export_row["current_count_qty"] = current_input
    return export_row


def _final_effective_qty(row: dict) -> int:
    for key in ("resolved_qty", "final_check_qty", "second_check_qty", "first_check_qty"):
        parsed = _parse_non_negative_int(row.get(key))
        if parsed is not None:
            return parsed
    return 0


def _front_inventory_final_count_map(session: dict) -> dict[str, int]:
    final_counts: dict[str, int] = {}
    for row in session.get("rows", []):
        if not isinstance(row, dict) or not _is_visible_inventory_row(row):
            continue
        part_number = _normalize_part_number(row.get("part_number"))
        if not part_number:
            continue
        final_counts[part_number] = _final_effective_qty(row)
    return final_counts


def _resolve_front_inventory_insight_template_path() -> Path | None:
    for candidate in (FRONT_INVENTORY_INSIGHT_TEMPLATE_OVERRIDE_PATH, FRONT_INVENTORY_INSIGHT_TEMPLATE_PATH):
        if candidate.exists():
            return candidate
    return None


def _front_inventory_insight_columns(sheet) -> tuple[int, int]:
    header_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = _build_header_map(header_values)
    part_index = (
        _find_header_index(header_map, ("alkatr", "szam"))
        or _find_header_index(header_map, ("cikk", "szam"))
        or _find_header_index(header_map, ("part", "number"))
    )
    if part_index is None:
        part_index = 0
    qty_index = (
        _find_header_index(header_map, ("darab",))
        or _find_header_index(header_map, ("menny",))
        or _find_header_index(header_map, ("qty",))
        or _find_header_index(header_map, ("quantity",))
    )
    if qty_index is None:
        qty_index = part_index + 1
    return part_index, qty_index


def _build_front_inventory_insight_ahk_script(quantities: list[int]) -> str:
    values_text = "\n".join(str(max(0, int(quantity or 0))) for quantity in quantities)
    return f"""; AutoHotkey v2.0+
#SingleInstance Force
Persistent

stepDelay := 220
startupDelay := 420
valuesText :=
(
{values_text}
)
isRunning := false

Esc::
{{
    ExitApp
}}

+Space::
{{
    global valuesText, stepDelay, startupDelay, isRunning
    if (isRunning)
        return
    isRunning := true

    while GetKeyState("Shift", "P") or GetKeyState("Space", "P")
    {{
        Sleep 40
    }}
    Sleep startupDelay

    values := StrSplit(valuesText, "`n", "`r")
    for _, quantity in values
    {{
        quantity := Trim(quantity)
        if (quantity = "")
            continue
        if GetKeyState("Esc", "P")
            ExitApp

        A_Clipboard := quantity
        Send "^v"
        Sleep stepDelay

        if GetKeyState("Esc", "P")
            ExitApp
        Send "{{Tab 2}}"
        Sleep stepDelay
    }}

    ExitApp
}}
"""


def _fill_finalize_workbook(workbook: Workbook, report_rows: list[dict]) -> None:
    summary_sheet = workbook.active
    summary_sheet.title = "Vegleges darabszamok"
    summary_sheet.append(["Alkatresz szam", "Darabszam"])
    for item in report_rows:
        summary_sheet.append([item["part_number"], item["final_count_qty"]])

    compare_sheet = workbook.create_sheet("Rendszer vs szamolt")
    compare_sheet.append(["Alkatresz szam", "Leiras", "Szamolt darabszam", "Raktari darabszam"])
    for item in report_rows:
        compare_sheet.append([item["part_number"], item["description"], item["final_count_qty"], item["stock_qty"]])

    rounds_sheet = workbook.create_sheet("Szamlalasok")
    rounds_sheet.append(["Alkatresz szam", "Leiras", "Elso szamlalas", "Masodik szamlalas", "Vegso szamlalas", "Raktari darabszam"])
    for item in report_rows:
        rounds_sheet.append(
            [
                item["part_number"],
                item["description"],
                _excel_cell_int(item.get("first_check_qty")),
                _excel_cell_int(item.get("second_check_qty")),
                item["final_count_qty"],
                item["stock_qty"],
            ]
        )

    for sheet, widths in (
        (summary_sheet, (28, 16)),
        (compare_sheet, (28, 56, 18, 18)),
        (rounds_sheet, (28, 56, 18, 18, 18, 18)),
    ):
        for width, column in zip(widths, ("A", "B", "C", "D", "E", "F")):
            sheet.column_dimensions[column].width = width


def build_front_inventory_view_model(session: dict, selected_category: str = "", sort_mode: str = "default") -> dict:
    finalized = str(session.get("phase")) == "finalized"
    source_rows = session.get("rows", [])
    for row in source_rows:
        if not str(row.get("color", "")).strip():
            row["color"] = _extract_color_value(row.get("description", ""), "")
        if not str(row.get("model", "")).strip():
            row["model"] = _extract_model_value(row.get("description", ""), row.get("part_number", ""))
        row["color"] = _normalize_inventory_row_color(
            str(row.get("model", "")).strip(),
            str(row.get("color", "")).strip(),
            str(row.get("description", "")).strip(),
        )
        row["color_label"] = _inventory_color_label(
            str(row.get("model", "")).strip(),
            str(row.get("color", "")).strip(),
            bool(row.get("is_serial")),
        )
    active_rows = _active_rows_for_phase(session, int(session.get("phase", 0) or 0)) if not finalized else [row for row in source_rows if _is_visible_inventory_row(row)]
    active_rows = list(active_rows)

    serial_rows = [row for row in active_rows if row.get("is_serial")]
    custom_rows = [row for row in active_rows if not row.get("is_serial")]

    categories: list[dict] = []
    categories.append({"key": "all", "label": "Összes", "count": len(serial_rows), "complete": bool(serial_rows) and all(str(row.get("input_qty", "")).strip() for row in serial_rows)})
    categories.append({"key": "egyedi", "label": "Egyedi", "count": len(custom_rows), "complete": bool(custom_rows) and all(str(row.get("input_qty", "")).strip() for row in custom_rows)})

    size_buckets: dict[str, list[dict]] = {}
    for row in serial_rows:
        category_key = str(row.get("category", "")).strip()
        if category_key:
            size_buckets.setdefault(category_key, []).append(row)

    for size_key in sorted(size_buckets, key=_inventory_category_sort_key):
        bucket_rows = size_buckets[size_key]
        categories.append({"key": size_key, "label": size_key, "count": len(bucket_rows), "complete": bool(bucket_rows) and all(str(row.get("input_qty", "")).strip() for row in bucket_rows)})

    allowed_keys = {item["key"] for item in categories if item["count"] > 0} | {"all", "egyedi"}
    active_category = selected_category if selected_category in allowed_keys else ("all" if serial_rows else ("egyedi" if custom_rows else "all"))

    if active_category == "all":
        visible_rows = serial_rows
    elif active_category == "egyedi":
        visible_rows = custom_rows
    else:
        visible_rows = [row for row in serial_rows if row.get("category") == active_category]

    active_sort = _normalize_inventory_sort_mode(sort_mode)
    if active_sort != "default":
        reverse = active_sort.endswith("_desc")
        base_sort = active_sort[:-5] if reverse else active_sort
        visible_rows = sorted(visible_rows, key=lambda row: _inventory_sort_key(row, base_sort, finalized), reverse=reverse)

    finalized_rows = _finalized_rows(session) if finalized else []

    return {
        "selected_category": active_category,
        "sort_mode": active_sort,
        "categories": categories,
        "visible_rows": visible_rows,
        "active_row_count": len(active_rows),
        "serial_row_count": len(serial_rows),
        "custom_row_count": len(custom_rows),
        "finalized": finalized,
        "finalized_rows": finalized_rows,
    }


def _finalized_rows(session: dict) -> list[dict]:
    rows = []
    for row in session.get("rows", []):
        if not _is_visible_inventory_row(row):
            continue
        resolved_qty = row.get("resolved_qty")
        if resolved_qty is None:
            continue
        model = str(row.get("model", "")).strip() or _extract_model_value(row.get("description", ""), row.get("part_number", ""))
        color = str(row.get("color", "")).strip() or _extract_color_value(row.get("description", ""), "")
        rows.append(
            {
                "part_number": row.get("part_number", ""),
                "description": row.get("description", ""),
                "model": model,
                "color": color,
                "color_label": _inventory_color_label(model, color, bool(row.get("is_serial"))),
                "counted_qty": int(resolved_qty or 0),
                "size": row.get("size", ""),
                "category": row.get("category", ""),
                "is_serial": bool(row.get("is_serial")),
            }
        )
    rows.sort(key=lambda row: (_row_category_sort_key(row), row.get("part_number", "")))
    return rows


def _active_rows_for_phase(session: dict, phase: int) -> list[dict]:
    if str(session.get("phase")) == "finalized":
        return []
    rows = []
    for row in session.get("rows", []):
        if not _is_visible_inventory_row(row):
            continue
        if row.get("status") == "resolved":
            continue
        if int(row.get("review_level", 0) or 0) == phase:
            rows.append(row)
    rows.sort(key=_row_sort_key)
    return rows


def _find_row(session: dict, row_id: str) -> dict | None:
    target = str(row_id or "").strip().upper()
    for row in session.get("rows", []):
        if str(row.get("row_id", "")).strip().upper() == target:
            return row
    return None


def _touch_session(session: dict) -> None:
    session["updated_at"] = datetime.now().isoformat(timespec="seconds")


def _read_stock_rows(file_name: str, payload: bytes, serial_sizes: set[str] | None = None) -> list[dict]:
    rows = _read_rows(file_name, payload)
    if not rows:
        raise ValueError("A front készletfájl üres.")

    header_map = _build_header_map(rows[0])
    part_index = _find_header_index(header_map, ("alkatr", "szam"))
    desc_index = _find_header_index(header_map, ("alkatr", "leiras"))
    qty_index = _find_header_index(header_map, ("rend", "all", "rakt", "keszl")) or _find_header_index(header_map, ("rend", "all"))
    color_index = _find_header_index(header_map, ("szin", "desc")) or _find_header_index(header_map, ("szin",))

    if None in {part_index, desc_index, qty_index}:
        raise ValueError("A front készletfájlban kell alkatrészszám, leírás és készlet oszlop.")

    items: list[dict] = []
    for row in rows[1:]:
        if max(part_index, desc_index, qty_index) >= len(row):
            continue
        part_number = _normalize_part_number(row[part_index])
        if not part_number:
            continue
        description = str(row[desc_index] or "").strip()
        quantity = _parse_non_negative_int(row[qty_index])
        color_value = row[color_index] if color_index is not None and color_index < len(row) else ""
        if quantity is None or quantity <= 0:
            continue
        if not _looks_like_front(part_number, description):
            continue
        model = _extract_model_value(description, part_number)
        color = _normalize_inventory_row_color(model, _extract_color_value(description, color_value), description)
        size = _extract_front_size(description, part_number)
        is_serial = bool(serial_sizes) and size in serial_sizes
        is_glass = _is_glass_front(part_number, description)
        if (not is_glass) and _is_excluded_inventory_row(model, color, is_serial):
            continue
        items.append(
            {
                "part_number": part_number,
                "description": description,
                "model": model,
                "color": color,
                "size": size,
                "is_serial": is_serial,
                "is_glass": is_glass,
                "quantity": quantity,
            }
        )

    if not items:
        raise ValueError("A front készletfájlban nem találtam raktáron lévő front sorokat.")
    return items


def _looks_like_front(part_number: str, description: str) -> bool:
    folded_description = _fold_text(description)
    return (
        str(part_number or "").strip().upper().startswith("NFA_")
        or str(part_number or "").strip().upper().startswith("NFAU_")
        or str(part_number or "").strip().upper().startswith("NFAH_")
        or str(part_number or "").strip().upper().startswith("NFAL_")
        or "folias fr" in folded_description
        or "front" in folded_description
        or "fo fr uv" in folded_description
        or "tak.s" in str(description or "").lower()
        or "sarok tak" in folded_description
    )


def _load_serial_sizes() -> set[str]:
    result: set[str] = set()
    for path, field_name in (
        (SERIAL_SIZES_DATA_PATH, "sizes"),
        (NETTFRONT_TRANSLATIONS_DATA_PATH, "standard_sizes"),
    ):
        if not path.exists():
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        values = payload.get(field_name, []) if isinstance(payload, dict) else []
        result.update(_normalize_front_size(value) for value in values)
    return {value for value in result if value}


def _repair_session(session: dict, session_path: Path | None = None) -> bool:
    changed = False
    serial_sizes = _load_serial_sizes()
    serial_size_list = sorted(serial_sizes, key=_size_sort_key)
    if session.get("serial_sizes") != serial_size_list:
        session["serial_sizes"] = serial_size_list
        changed = True

    for row in session.get("rows", []):
        if not isinstance(row, dict):
            continue
        description = str(row.get("description", ""))
        part_number = str(row.get("part_number", ""))
        size = _extract_front_size(description, part_number)
        is_serial = bool(size) and size in serial_sizes
        is_glass = _is_glass_front(part_number, description)
        model = str(row.get("model", "")).strip() or _extract_model_value(description, part_number)
        color = _normalize_inventory_row_color(model, str(row.get("color", "")).strip(), description)
        color_label = _inventory_color_label(model, color, is_serial)
        category = _inventory_category_key(size, is_serial, is_glass)

        for key, value in (
            ("size", size),
            ("is_serial", is_serial),
            ("is_glass", is_glass),
            ("category", category),
            ("model", model),
            ("color", color),
            ("color_label", color_label),
        ):
            if row.get(key) != value:
                row[key] = value
                changed = True
    if _merge_saved_stock_rows(session, session_path, serial_sizes):
        changed = True
    return changed


def _merge_saved_stock_rows(session: dict, session_path: Path | None, serial_sizes: set[str]) -> bool:
    if session_path is None:
        return False
    stock_path = _find_saved_stock_upload_path(session_path.parent)
    if stock_path is None or not stock_path.exists():
        return False

    try:
        stock_rows = _read_stock_rows(stock_path.name, stock_path.read_bytes(), serial_sizes)
    except Exception:
        return False

    current_rows = [row for row in session.get("rows", []) if isinstance(row, dict)]
    existing_parts = {str(row.get("part_number", "")).strip().upper() for row in current_rows}
    added = False
    for item in stock_rows:
        part_number = str(item.get("part_number", "")).strip().upper()
        if not part_number or part_number in existing_parts:
            continue
        current_rows.append(_build_inventory_session_row(item, serial_sizes))
        existing_parts.add(part_number)
        added = True

    if added:
        current_rows.sort(key=_row_sort_key)
        session["rows"] = current_rows
    return added


def _find_saved_stock_upload_path(runtime_dir: Path) -> Path | None:
    for candidate in sorted(runtime_dir.glob("latest-stock.*")):
        if candidate.suffix.lower() in FRONT_INVENTORY_ALLOWED_EXTENSIONS:
            return candidate
    return None


def _build_inventory_session_row(item: dict, serial_sizes: set[str]) -> dict:
    description = str(item.get("description", "")).strip()
    part_number = str(item.get("part_number", "")).strip()
    model = str(item.get("model", "")).strip() or _extract_model_value(description, part_number)
    color = _normalize_inventory_row_color(model, str(item.get("color", "")).strip(), description)
    size = str(item.get("size", "")).strip() or _extract_front_size(description, part_number)
    is_serial = bool(size) and size in serial_sizes
    is_glass = bool(item.get("is_glass")) or _is_glass_front(part_number, description)
    return {
        "row_id": part_number,
        "part_number": part_number,
        "description": description,
        "model": model,
        "color": color,
        "color_label": _inventory_color_label(model, color, is_serial),
        "stock_qty": int(item.get("quantity", 0) or 0),
        "size": size,
        "category": _inventory_category_key(size, is_serial, is_glass),
        "is_serial": is_serial,
        "is_glass": is_glass,
        "input_qty": "",
        "resolved_qty": None,
        "first_check_qty": None,
        "second_check_qty": None,
        "final_check_qty": None,
        "review_level": 0,
        "status": "pending",
    }


def _read_rows(file_name: str, payload: bytes) -> list[list[object]]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise RuntimeError("Az Excel feldolgozáshoz hiányzik az openpyxl csomag.")
        workbook = load_workbook(io.BytesIO(payload), data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    if suffix == ".csv":
        text = _decode_csv_bytes(payload)
        return [list(row) for row in csv.reader(io.StringIO(text))]
    raise ValueError("Csak XLSX, XLSM vagy CSV fájlokat tudok feldolgozni.")


def _decode_csv_bytes(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="ignore")


def _build_header_map(header_row: list[object]) -> dict[int, str]:
    return {index: _normalize_header(value) for index, value in enumerate(header_row)}


def _find_header_index(header_map: dict[int, str], terms: tuple[str, ...]) -> int | None:
    for index, normalized in header_map.items():
        if all(term in normalized for term in terms):
            return index
    return None


def _normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", _fold_text(value))


def _normalize_part_number(value: object) -> str:
    return str(value or "").strip().upper()


def _parse_non_negative_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except Exception:
            return None
        if number < 0:
            return None
        return int(round(number))
    text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    if number < 0:
        return None
    return int(round(number))


def _extract_front_size(description: str, part_number: str) -> str:
    for source in (str(description or ""), str(part_number or "")):
        match = re.search(r"(\d{2,4}\s*x\s*\d{2,4})\s*x\s*\d{1,2}\b", source, flags=re.IGNORECASE)
        if match:
            return _normalize_front_size(match.group(1))
        match = re.search(r"(\d{2,4}\s*x\s*\d{2,4})\b", source, flags=re.IGNORECASE)
        if match:
            return _normalize_front_size(match.group(1))
    return ""


def _extract_color_value(description: object, explicit_color: object) -> str:
    clean_color = str(explicit_color or "").strip()
    if clean_color:
        return clean_color

    text = str(description or "").strip()
    if not text:
        return ""

    text = re.sub(r"^Fóliás fr\.?", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\b\d{2,4}\s*x\s*\d{2,4}\s*x\s*\d{1,2}\b", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\b\d{2,4}\s*x\s*\d{2,4}\b", "", text, flags=re.IGNORECASE).strip()
    parts = text.split()
    if len(parts) <= 1:
        return text
    return " ".join(parts[1:]).strip()


def _extract_model_value(description: object, part_number: object = "") -> str:
    text = str(description or "").strip()
    if text:
        text = re.sub(r"^F[oó]li[aá]s\s*fr\.?", "", text, flags=re.IGNORECASE).strip()
        if text:
            return text.split()[0].strip()

    tokens = str(part_number or "").strip().upper().split("_")
    if len(tokens) > 1:
        return {
            "ANT": "Antónia",
            "LU": "Laura",
            "ZI": "Zille",
        }.get(tokens[1], "")
    return ""


def _compose_color_label(model: str, color: str) -> str:
    clean_model = str(model or "").strip()
    clean_color = str(color or "").strip()
    if clean_model and clean_color and not _fold_text(clean_color).startswith(_fold_text(clean_model)):
        return f"{clean_model} {clean_color}".strip()
    return clean_color or clean_model


def _inventory_color_label(model: str, color: str, is_serial: bool) -> str:
    clean_color = str(color or "").strip()
    if not is_serial:
        return clean_color
    return _compose_color_label(model, clean_color)


def _normalize_inventory_row_color(model: str, color: str, description: str = "") -> str:
    clean_color = _cleanup_inventory_color_text(color)
    if not clean_color:
        clean_color = _cleanup_inventory_color_text(_extract_color_value(description, ""))
    if not clean_color:
        return ""
    candidates = [str(model or "").strip(), "Antónia", "Laura", "Zille"]
    for candidate in candidates:
        if not candidate:
            continue
        clean_color = re.sub(rf"^{re.escape(candidate)}\b[\s.:/-]*", "", clean_color, flags=re.IGNORECASE).strip()
    return clean_color


def _cleanup_inventory_color_text(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r"^(?:EGYEDI|ROSSZ)\b[\s.:/-]*", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"^F[oó]li[aá]s\s*fr\.?", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\b\d{2,4}\s*x\s*\d{2,4}\s*x\s*\d{1,2}\b", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\b\d{2,4}\s*x\s*\d{2,4}\b", "", text, flags=re.IGNORECASE).strip()
    return re.sub(r"\s+", " ", text).strip(" -")


def _normalized_inventory_color(value: object) -> str:
    normalized = _fold_text(value).replace(".", " ").replace("-", " ").replace("_", " ")
    return re.sub(r"\s+", " ", normalized).strip()


def _is_excluded_inventory_color(model: object, color: object) -> bool:
    model_key = _fold_text(model)
    color_key = _normalized_inventory_color(color)
    if not color_key or color_key.startswith("sm "):
        return False

    if color_key.startswith("matt "):
        if any(token in color_key for token in ("beige", "capuccino", "cappuccino", "feher", "grafit", "szurke")):
            return True
        if "provance" in color_key and model_key == "laura":
            return True
    return False


def _is_excluded_inventory_row(model: object, color: object, is_serial: bool) -> bool:
    return bool(is_serial) and _is_excluded_inventory_color(model, color)


def _is_visible_inventory_row(row: dict) -> bool:
    model = str(row.get("model", "")).strip() or _extract_model_value(row.get("description", ""), row.get("part_number", ""))
    color = _normalize_inventory_row_color(model, str(row.get("color", "")).strip(), str(row.get("description", "")))
    is_serial = bool(row.get("is_serial"))
    is_glass = bool(row.get("is_glass")) or _is_glass_front(row.get("part_number", ""), row.get("description", ""))
    size = _normalize_front_size(row.get("size", "")) or _extract_front_size(row.get("description", ""), row.get("part_number", ""))
    row["model"] = model
    row["color"] = color
    row["is_glass"] = is_glass
    row["size"] = size
    row["color_label"] = _inventory_color_label(model, color, is_serial)
    return not (not is_glass and _is_excluded_inventory_row(model, color, is_serial))


def _normalize_inventory_sort_mode(value: object) -> str:
    clean = str(value or "").strip().lower()
    allowed = {"default", "color", "color_desc", "description", "description_desc", "count", "count_desc"}
    return clean if clean in allowed else "default"


def _inventory_sort_key(row: dict, mode: str, finalized: bool) -> tuple:
    if mode == "description":
        return (_fold_text(row.get("description", "")), _fold_text(row.get("color_label", row.get("color", ""))), str(row.get("part_number", "")))
    if mode == "count":
        if finalized:
            count_value = int(row.get("counted_qty", 0) or 0)
        else:
            parsed = _row_input_value(row)
            count_value = parsed if parsed is not None else -1
        return (count_value, _fold_text(row.get("description", "")), str(row.get("part_number", "")))
    return (_fold_text(row.get("color_label", row.get("color", ""))), _fold_text(row.get("description", "")), str(row.get("part_number", "")))


def _normalize_front_size(value: object) -> str:
    clean_value = re.sub(r"\s+", "", str(value or "")).lower()
    match = re.match(r"^(\d{2,4})x(\d{2,4})$", clean_value)
    if not match:
        return ""
    return f"{match.group(1)}x{match.group(2)}"


def _size_sort_key(value: str) -> tuple[int, int, str]:
    match = re.match(r"^(\d{2,4})x(\d{2,4})$", str(value or "").strip().lower())
    if not match:
        return (9999, 9999, str(value or ""))
    return (int(match.group(1)), int(match.group(2)), str(value or ""))


def _inventory_category_key(size: str, is_serial: bool, is_glass: bool) -> str:
    if not is_serial:
        return "egyedi"
    clean_size = str(size or "").strip()
    if is_glass and clean_size:
        return f"Üveges - {clean_size}"
    return clean_size


def _inventory_category_sort_key(value: str) -> tuple[int, int, int, str]:
    clean_value = str(value or "").strip()
    if clean_value == "egyedi":
        return (2, 9999, 9999, clean_value)
    if clean_value.lower().startswith("üveges - ".lower()):
        base_size = clean_value.split("-", 1)[1].strip()
        w, h, _ = _size_sort_key(base_size)
        return (1, w, h, clean_value)
    w, h, _ = _size_sort_key(clean_value)
    if w != 9999 or h != 9999:
        return (0, w, h, clean_value)
    return (3, 9999, 9999, clean_value)


def _is_glass_front(part_number: object, description: object = "") -> bool:
    clean_part = str(part_number or "").strip().upper()
    if clean_part.startswith("NFAU_"):
        return True
    folded_description = _fold_text(description)
    return "fo fr uv" in folded_description or "fo fr.uv" in folded_description or "fo früv" in folded_description


def _row_category_sort_key(row: dict) -> tuple[int, tuple[int, int, str] | tuple[int, int, str], str]:
    if row.get("is_serial"):
        return (0, _inventory_category_sort_key(str(row.get("category", ""))), str(row.get("part_number", "")))
    return (1, (9999, 9999, "egyedi"), str(row.get("part_number", "")))


def _row_sort_key(row: dict) -> tuple[int, tuple[int, int, str], str]:
    return _row_category_sort_key(row)


def _row_input_value(row: dict) -> int | None:
    return _parse_non_negative_int(row.get("input_qty"))


def _excel_cell_int(value: object) -> int | str:
    parsed = _parse_non_negative_int(value)
    return parsed if parsed is not None else ""


def _fold_text(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(char for char in normalized if not unicodedata.combining(char)).casefold().strip()
