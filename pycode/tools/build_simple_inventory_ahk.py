from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_ROOT = REPO_ROOT / "runtime" / "front-leltar" / "ahs-riport"


def _normalize_header(value: object) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def _parse_quantity(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return str(int(value))
    if isinstance(value, (int, float)):
        return str(int(value))
    text = str(value).strip().replace(" ", "")
    if not text:
        return ""
    text = text.replace(",", ".")
    try:
        return str(int(float(text)))
    except Exception:
        return text


def _find_columns(sheet) -> tuple[int, int]:
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized = [_normalize_header(value) for value in headers]
    part_index = 0
    qty_index = 1
    for index, header in enumerate(normalized):
        if "alkatr" in header and "szam" in header:
            part_index = index
        if "menny" in header or "darab" in header or "qty" in header:
            qty_index = index
    if qty_index == part_index:
        qty_index = part_index + 1
    return part_index, qty_index


def _load_count_map(source_path: Path) -> dict[str, str]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    part_index, qty_index = _find_columns(sheet)
    count_map: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if max(part_index, qty_index) >= len(row):
            continue
        part_number = str(row[part_index] or "").strip().upper()
        if not part_number:
            continue
        count_map[part_number] = _parse_quantity(row[qty_index])
    return count_map


def _build_ahk_script(values_file_name: str) -> str:
    return f"""; AutoHotkey v2.0+
#SingleInstance Force
Persistent

stepDelay := 220
tabDelay := 40
startupDelay := 420
valuesFile := A_ScriptDir "\\{values_file_name}"
backspaceToken := "__BACKSPACE__"
isRunning := false

Esc::
{{
    ExitApp
}}

+Space::
{{
    global valuesFile, backspaceToken, stepDelay, tabDelay, startupDelay, isRunning
    if (isRunning)
        return
    isRunning := true

    while GetKeyState("Shift", "P") or GetKeyState("Space", "P")
    {{
        Sleep 40
    }}
    Sleep startupDelay

    valuesText := FileRead(valuesFile, "UTF-8")
    valuesText := StrReplace(valuesText, "`r", "")
    values := StrSplit(valuesText, "`n")
    for _, quantity in values
    {{
        quantity := Trim(quantity)
        if GetKeyState("Esc", "P")
            ExitApp

        if (quantity = backspaceToken)
            Send("{{Backspace}}")
        else if (quantity != "")
            SendText(quantity)
        Sleep stepDelay
        Send("{{Tab}}")
        Sleep tabDelay
        Send("{{Tab}}")
        Sleep stepDelay
    }}

    ExitApp
}}
"""


def build_package(source_path: Path, output_dir: Path, counts_path: Path | None = None) -> dict[str, Path | int]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    part_index, qty_index = _find_columns(sheet)
    count_map = _load_count_map(counts_path) if counts_path else {}

    ordered_quantities: list[str] = []
    row_count = 0
    matched_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if part_index >= len(row):
            continue
        part_number = str(row[part_index] or "").strip().upper()
        if not part_number:
            continue
        if count_map:
            quantity = count_map.get(part_number, "__BACKSPACE__")
            if part_number in count_map:
                matched_count += 1
        else:
            if qty_index >= len(row):
                continue
            quantity = _parse_quantity(row[qty_index]) or "__BACKSPACE__"
        ordered_quantities.append(quantity)
        row_count += 1

    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    workbook_path = output_dir / f"kasmir-provance-bevetelezes-{stamp}.xlsx"
    values_path = output_dir / f"kasmir-provance-bevetelezes-{stamp}.txt"
    ahk_path = output_dir / f"kasmir-provance-bevetelezes-{stamp}.ahk"
    ahs_path = output_dir / f"kasmir-provance-bevetelezes-{stamp}.ahs"
    meta_path = output_dir / f"kasmir-provance-bevetelezes-{stamp}.json"

    latest_workbook_path = output_dir / "kasmir-provance-bevetelezes-latest.xlsx"
    latest_values_path = output_dir / "kasmir-provance-bevetelezes-latest.txt"
    latest_ahk_path = output_dir / "kasmir-provance-bevetelezes-latest.ahk"
    latest_ahs_path = output_dir / "kasmir-provance-bevetelezes-latest.ahs"
    latest_meta_path = output_dir / "kasmir-provance-bevetelezes-latest.json"

    shutil.copy2(source_path, workbook_path)
    values_path.write_text("\n".join(ordered_quantities), encoding="utf-8")
    script_body = _build_ahk_script(values_path.name)
    ahk_path.write_text(script_body, encoding="utf-8")
    ahs_path.write_text(script_body, encoding="utf-8")
    meta_path.write_text(
        json.dumps(
            {
                "source": str(source_path),
                "counts_source": str(counts_path) if counts_path else "",
                "workbook": str(workbook_path),
                "values": str(values_path),
                "ahk": str(ahk_path),
                "ahs": str(ahs_path),
                "rows": row_count,
                "matched_rows": matched_count,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    shutil.copy2(workbook_path, latest_workbook_path)
    shutil.copy2(values_path, latest_values_path)
    shutil.copy2(ahk_path, latest_ahk_path)
    shutil.copy2(ahs_path, latest_ahs_path)
    shutil.copy2(meta_path, latest_meta_path)

    return {
        "workbook": workbook_path,
        "values": values_path,
        "ahk": ahk_path,
        "ahs": ahs_path,
        "meta": meta_path,
        "rows": row_count,
        "matched_rows": matched_count,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--counts", type=Path, default=None)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_ROOT)
    args = parser.parse_args()
    result = build_package(args.source, args.output_dir, args.counts)
    print("WORKBOOK", result["workbook"])
    print("VALUES", result["values"])
    print("AHK", result["ahk"])
    print("AHS", result["ahs"])
    print("META", result["meta"])
    print("ROWS", result["rows"])
    print("MATCHED_ROWS", result["matched_rows"])


if __name__ == "__main__":
    main()
