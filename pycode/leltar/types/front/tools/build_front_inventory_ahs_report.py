"""Build front inventory ahs report helpers for the inventory package."""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[5]
DOWNLOADS_DIR = Path.home() / "Downloads"
OUTPUT_ROOT = REPO_ROOT / "runtime" / "front-leltar" / "ahs-riport"
PYCODE_DIR = REPO_ROOT / "pycode"
if str(PYCODE_DIR) not in sys.path:
    sys.path.insert(0, str(PYCODE_DIR))

def _normalize_header(value: object) -> str:
    """Normalize a worksheet header for tolerant column matching."""
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def _normalize_part_number(value: object) -> str:
    """Normalize part numbers used as workbook lookup keys."""
    return str(value or "").strip().upper()


def _parse_count(value: object) -> int:
    """Parse a counted quantity, returning zero for blank or invalid values."""
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return 0
    text = text.replace(",", ".")
    try:
        return int(float(text))
    except Exception:
        return 0


def _pick_latest(pattern: str) -> Path:
    """Return the newest downloaded file matching a glob pattern."""
    candidates = [p for p in DOWNLOADS_DIR.glob(pattern) if p.is_file() and not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"Nincs találat erre: {pattern}")
    candidates.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return candidates[0]


def _find_template_columns(sheet) -> tuple[int, int]:
    """Locate template part-number and quantity columns."""
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized = [_normalize_header(value) for value in headers]
    part_index = 0
    qty_index = 1
    for index, header in enumerate(normalized):
        if "alkatr" in header and "szam" in header:
            part_index = index
        if ("szamolt" in header and "menny" in header) or "darab" in header or "menny" in header:
            qty_index = index
    if qty_index == part_index:
        qty_index = part_index + 1
    return part_index, qty_index


def _find_count_columns(sheet) -> tuple[int, int]:
    """Locate source workbook part-number and counted-quantity columns."""
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized = [_normalize_header(value) for value in headers]
    part_index = 0
    qty_index = 1
    for index, header in enumerate(normalized):
        if "alkatr" in header and "szam" in header:
            part_index = index
        if "szamlalas" in header or "darab" in header or "menny" in header:
            qty_index = index
    if qty_index == part_index:
        qty_index = part_index + 1
    return part_index, qty_index


def _build_ahk_script(values_file_name: str) -> str:
    """Build the AutoHotkey script that pastes count values."""
    return f"""; AutoHotkey v2.0+
#SingleInstance Force
Persistent

stepDelay := 220
tabDelay := 40
startupDelay := 420
valuesFile := A_ScriptDir "\\{values_file_name}"
isRunning := false

Esc::
{{
    ExitApp
}}

+Space::
{{
    global valuesFile, stepDelay, tabDelay, startupDelay, isRunning
    if (isRunning)
        return
    isRunning := true

    while GetKeyState("Shift", "P") or GetKeyState("Space", "P")
    {{
        Sleep 40
    }}
    Sleep startupDelay

    valuesText := FileRead(valuesFile, "UTF-8")
    valuesText := StrReplace(valuesText, "`r", "")
    values := StrSplit(valuesText, "`n")
    for _, quantity in values
    {{
        quantity := Trim(quantity)
        if (quantity = "")
            continue
        if GetKeyState("Esc", "P")
            ExitApp

        SendText(quantity)
        Sleep stepDelay
        Send("{{Tab}}")
        Sleep tabDelay
        Send("{{Tab}}")
        Sleep stepDelay
    }}

    ExitApp
}}
"""


def build_report(template_path: Path, counts_path: Path, output_dir: Path) -> dict[str, object]:
    """Build the filled AHS workbook and companion import artifacts."""
    template_wb = load_workbook(template_path)
    template_ws = template_wb[template_wb.sheetnames[0]]
    count_wb = load_workbook(counts_path, read_only=True, data_only=True)
    count_ws = count_wb[count_wb.sheetnames[0]]

    template_part_col, template_qty_col = _find_template_columns(template_ws)
    count_part_col, count_qty_col = _find_count_columns(count_ws)

    count_map: dict[str, int] = {}
    for row in count_ws.iter_rows(min_row=2, values_only=True):
        if max(count_part_col, count_qty_col) >= len(row):
            continue
        part_number = _normalize_part_number(row[count_part_col])
        if not part_number:
            continue
        count_map[part_number] = _parse_count(row[count_qty_col])

    ordered_quantities: list[str] = []
    matched_parts: set[str] = set()
    row_count = 0
    for row_index in range(2, template_ws.max_row + 1):
        part_number = _normalize_part_number(template_ws.cell(row=row_index, column=template_part_col + 1).value)
        if not part_number:
            continue
        quantity = int(count_map.get(part_number, 0))
        template_ws.cell(row=row_index, column=template_qty_col + 1).value = quantity
        ordered_quantities.append(str(quantity))
        matched_parts.add(part_number)
        row_count += 1

    missing_parts = [(part_number, count_map[part_number]) for part_number in count_map if part_number not in matched_parts]
    for part_number, quantity in missing_parts:
        template_ws.append([part_number, quantity])
        ordered_quantities.append(str(int(quantity)))
        row_count += 1

    missing_sheet = template_wb.create_sheet("Nincs a mintaban")
    missing_sheet.append(["Alkatresz szam", "Szamolt mennyiseg"])
    for part_number, quantity in missing_parts:
        missing_sheet.append([part_number, int(quantity)])

    meta_sheet = template_wb.create_sheet("Osszesites")
    meta_sheet.append(["Mutato", "Ertek"])
    meta_sheet.append(["Minta fajl", template_path.name])
    meta_sheet.append(["Szamlalas fajl", counts_path.name])
    meta_sheet.append(["Mintaban talalt sorok", len(matched_parts)])
    meta_sheet.append(["Minta utan hozzafuzott sorok", len(missing_parts)])
    meta_sheet.append(["Vegso sorok szama", row_count])

    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    workbook_path = output_dir / f"front-leltar-ahs-riport-{stamp}.xlsx"
    ahk_path = output_dir / f"front-leltar-ahs-riport-{stamp}.ahk"
    ahs_path = output_dir / f"front-leltar-ahs-riport-{stamp}.ahs"
    values_path = output_dir / f"front-leltar-ahs-riport-{stamp}.txt"
    meta_path = output_dir / f"front-leltar-ahs-riport-{stamp}.json"
    latest_workbook_path = output_dir / "front-leltar-ahs-riport-latest.xlsx"
    latest_ahk_path = output_dir / "front-leltar-ahs-riport-latest.ahk"
    latest_ahs_path = output_dir / "front-leltar-ahs-riport-latest.ahs"
    latest_values_path = output_dir / "front-leltar-ahs-riport-latest.txt"
    latest_meta_path = output_dir / "front-leltar-ahs-riport-latest.json"

    template_wb.save(workbook_path)
    values_path.write_text("\n".join(ordered_quantities), encoding="utf-8")
    script_body = _build_ahk_script(values_path.name)
    ahk_path.write_text(script_body, encoding="utf-8")
    ahs_path.write_text(script_body, encoding="utf-8")
    meta_path.write_text(
        json.dumps(
            {
                "template": str(template_path),
                "counts": str(counts_path),
                "workbook": str(workbook_path),
                "ahk": str(ahk_path),
                "ahs": str(ahs_path),
                "values": str(values_path),
                "matched_parts": len(matched_parts),
                "missing_parts": len(missing_parts),
                "final_rows": row_count,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    shutil.copy2(workbook_path, latest_workbook_path)
    shutil.copy2(ahk_path, latest_ahk_path)
    shutil.copy2(ahs_path, latest_ahs_path)
    shutil.copy2(values_path, latest_values_path)
    shutil.copy2(meta_path, latest_meta_path)

    return {
        "workbook": workbook_path,
        "ahk": ahk_path,
        "ahs": ahs_path,
        "values": values_path,
        "meta": meta_path,
        "matched_parts": len(matched_parts),
        "missing_parts": len(missing_parts),
        "final_rows": row_count,
    }


def main() -> None:
    """Command-line entry point for the AHS report builder."""
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", type=Path, default=None)
    parser.add_argument("--counts", type=Path, default=None)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_ROOT)
    args = parser.parse_args()

    template_path = args.template or _pick_latest("front lelt*.xlsx")
    counts_path = args.counts or _pick_latest("Leltár számolás*.xlsx")
    result = build_report(template_path, counts_path, args.output_dir)
    print("WORKBOOK", result["workbook"])
    print("AHK", result["ahk"])
    print("AHS", result["ahs"])
    print("VALUES", result["values"])
    print("META", result["meta"])
    print("MATCHED_PARTS", result["matched_parts"])
    print("MISSING_PARTS", result["missing_parts"])
    print("FINAL_ROWS", result["final_rows"])


if __name__ == "__main__":
    main()
