"""Job persistence and artifact generation for NettFront order runs.

This module stores order suggestion rows, produces downloadable CSV and Excel
outputs, and keeps approval state in runtime metadata.
"""

from __future__ import annotations

import csv
import io
import json
import mimetypes
import uuid
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

from nettfront.engine import create_bundle_archive
from nettfront.tools.jobs import download_payload, read_job

from .config import order_runtime_dir
from .engine import (
    NettfrontOrderRow,
    calc_total_m2_from_rows,
    rows_to_approved_workbook,
    rows_to_suggestion_workbook,
)


def _format_eu_number(value: float, decimals: int = 2) -> str:
    """Format eu number values for display or export."""
    formatted = f"{value:,.{decimals}f}"
    return formatted.replace(",", "_").replace(".", ",").replace("_", ".")


def read_order_job(job_id: str) -> tuple[Path | None, dict | None]:
    """Read order job data."""
    return read_job(order_runtime_dir(), job_id)


def order_download_payload(job_id: str, artifact: str) -> tuple[bytes, str, str] | None:
    """Handle order download payload logic for the NettFront workflows."""
    _job_dir, metadata = read_order_job(job_id)
    if metadata is None:
        return None

    source_stock_file = str(metadata.get("source_stock_file", "")).strip()
    source_stock_name = str(metadata.get("source_stock_name", source_stock_file)).strip() or source_stock_file
    guessed_stock_type = mimetypes.guess_type(source_stock_name)[0] or "application/octet-stream"
    artifact_map = {
        "suggestion-xlsx": (
            "rendelesi-javaslat.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "rendelesi-javaslat.xlsx",
        ),
        "approved-xlsx": (
            metadata.get("approved_file", "rendeles-jovahagyott.xlsx"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata.get("approved_file", "rendeles-jovahagyott.xlsx"),
        ),
        "import-csv": (
            metadata.get("import_file", "rendeles_sima.csv"),
            "text/csv; charset=utf-8",
            metadata.get("import_file", "rendeles_sima.csv"),
        ),
        "source-stock": (
            source_stock_file,
            guessed_stock_type,
            source_stock_name,
        ),
        "bundle-zip": (
            metadata.get("bundle_name", "nettfront-rendeles-output.zip"),
            "application/zip",
            metadata.get("bundle_name", "nettfront-rendeles-output.zip"),
        ),
    }

    return download_payload(order_runtime_dir(), job_id, artifact, artifact_map)

def _order_safe_number(value) -> float:
    """Handle order safe number logic for the NettFront workflows."""
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return 0.0
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _order_parse_quantity_input(value: str) -> tuple[float, bool]:
    """Handle order parse quantity input logic for the NettFront workflows."""
    text = str(value or "").strip()
    if not text:
        return 0.0, True
    sanitized = text.replace(" ", "")
    if "," in sanitized and "." in sanitized:
        if sanitized.rfind(",") > sanitized.rfind("."):
            sanitized = sanitized.replace(".", "").replace(",", ".")
        else:
            sanitized = sanitized.replace(",", "")
    elif "," in sanitized:
        sanitized = sanitized.replace(",", ".")
    try:
        return max(0.0, float(sanitized)), True
    except ValueError:
        return 0.0, False


def _format_order_metric(value) -> str:
    """Format order metric values for display or export."""
    if value in (None, ""):
        return "—"
    raw = str(value).strip()
    if not raw:
        return "—"
    if not any(char.isdigit() for char in raw):
        return raw
    number = _order_safe_number(value)
    decimals = 0 if abs(number - round(number)) < 1e-9 else 2
    return _format_eu_number(number, decimals)


def _format_order_input_value(value) -> str:
    """Format order input value values for display or export."""
    number = _order_safe_number(value)
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    return f"{number:.2f}".rstrip("0").rstrip(".").replace(".", ",")


def _count_positive_order_rows(rows: list[NettfrontOrderRow]) -> int:
    """Handle count positive order rows logic for the NettFront workflows."""
    return sum(1 for row in rows if _order_safe_number(row.order_qty) > 0)


def _nettfront_order_row_to_dict(row: NettfrontOrderRow) -> dict:
    """Handle nettfront order row to dict logic for the NettFront workflows."""
    return {
        "row_id": row.row_id,
        "part_number": row.part_number,
        "description": row.description,
        "stock_unit": row.stock_unit,
        "current_stock": row.current_stock,
        "confirmed_demand": row.confirmed_demand,
        "open_procurement": row.open_procurement,
        "safe_stock": row.safe_stock,
        "capacity": row.capacity,
        "order_qty": row.order_qty,
        "color": row.color,
        "length": row.length,
        "width": row.width,
        "is_super_matt": row.is_super_matt,
    }


def _nettfront_order_row_from_dict(payload: dict) -> NettfrontOrderRow:
    """Handle nettfront order row from dict logic for the NettFront workflows."""
    return NettfrontOrderRow(
        row_id=str(payload.get("row_id", "")).strip(),
        part_number=str(payload.get("part_number", "")).strip(),
        description=str(payload.get("description", "")).strip(),
        stock_unit=payload.get("stock_unit"),
        current_stock=payload.get("current_stock"),
        confirmed_demand=payload.get("confirmed_demand"),
        open_procurement=payload.get("open_procurement"),
        safe_stock=payload.get("safe_stock"),
        capacity=payload.get("capacity"),
        order_qty=_order_safe_number(payload.get("order_qty")),
        color=str(payload.get("color", "")).strip(),
        length=_order_safe_number(payload.get("length")),
        width=_order_safe_number(payload.get("width")),
        is_super_matt=bool(payload.get("is_super_matt")),
    )


def _read_nettfront_order_rows(job_dir: Path) -> list[NettfrontOrderRow]:
    """Read nettfront order rows data."""
    rows_path = job_dir / "suggestions.json"
    if not rows_path.exists():
        return []
    try:
        payload = json.loads(rows_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    if not isinstance(payload, list):
        return []
    return [_nettfront_order_row_from_dict(item) for item in payload if isinstance(item, dict)]


def _write_nettfront_order_rows(job_dir: Path, rows: list[NettfrontOrderRow]) -> None:
    """Write nettfront order rows data."""
    payload = [_nettfront_order_row_to_dict(row) for row in rows]
    (job_dir / "suggestions.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _nettfront_order_quantity_text(value: float) -> str:
    """Handle nettfront order quantity text logic for the NettFront workflows."""
    number = _order_safe_number(value)
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _normalize_nettfront_part_number(value: object) -> str:
    """Normalize nettfront part number values."""
    text = str(value or "").strip().upper()
    return re.sub(r"\s+", "", text)


def _nettfront_parts_list_header_key(value: str) -> str:
    """Handle nettfront parts list header key logic for the NettFront workflows."""
    return re.sub(r"[^A-Z0-9]+", "", _normalize_nettfront_part_number(value))


def _nettfront_order_part_number_aliases(value: object) -> list[str]:
    """Handle nettfront order part number aliases logic for the NettFront workflows."""
    normalized = _normalize_nettfront_part_number(value)
    if not normalized:
        return []

    aliases = [normalized]
    for base_tag, secondary_tag, merged_tag in (("KAF", "KAFS", "KAFU"), ("PRA", "PRAS", "PRAU")):
        match = re.match(rf"^(NFA[^_]*_ANT)_{merged_tag}_(.+)$", normalized)
        if not match:
            continue
        base = match.group(1)
        suffix = match.group(2)
        aliases.extend(
            [
                f"{base}_{base_tag}_{suffix}",
                f"{base}_{secondary_tag}_{suffix}",
            ]
        )
        break

    unique_aliases: list[str] = []
    seen: set[str] = set()
    for alias in aliases:
        if alias in seen:
            continue
        seen.add(alias)
        unique_aliases.append(alias)
    return unique_aliases


def _nettfront_order_display_part_number(value: object) -> str:
    """Handle nettfront order display part number logic for the NettFront workflows."""
    aliases = _nettfront_order_part_number_aliases(value)
    if not aliases:
        return ""
    if len(aliases) >= 2 and aliases[0] != aliases[1]:
        return aliases[1]
    return aliases[0]


def _load_nettfront_parts_list_from_bytes(payload: bytes, file_name: str) -> list[str]:
    """Load nettfront parts list from bytes data."""
    file_name = str(file_name or "").strip().lower()
    values: list[str] = []

    if file_name.endswith((".xlsx", ".xlsm")):
        if load_workbook is None:
            raise ValueError("Az Excel feldolgozáshoz hiányzik az openpyxl csomag.")
        workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        worksheet = workbook.active
        for row in worksheet.iter_rows(values_only=True):
            first_value = None
            for cell in row:
                if cell not in (None, ""):
                    first_value = cell
                    break
            normalized = _normalize_nettfront_part_number(first_value)
            if normalized:
                values.append(normalized)
    elif file_name.endswith(".csv"):
        decoded = None
        for encoding in ("utf-8-sig", "cp1250", "cp1252", "latin-1"):
            try:
                decoded = payload.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise ValueError("A CSV fájl kódolását nem tudtam beolvasni.")
        for row in csv.reader(io.StringIO(decoded)):
            first_value = next((cell for cell in row if str(cell).strip()), "")
            normalized = _normalize_nettfront_part_number(first_value)
            if normalized:
                values.append(normalized)
    else:
        raise ValueError("A friss alkatrészlista csak XLSX, XLSM vagy CSV lehet.")

    unique_values: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not unique_values and _nettfront_parts_list_header_key(value) in {
            "ALKATRESZ",
            "ALKATRESZSZAM",
            "ALKATRSZAM",
            "CIKKSZAM",
            "PARTNUMBER",
            "PARTNUM",
        }:
            continue
        if value in seen:
            continue
        seen.add(value)
        unique_values.append(value)
    return unique_values


def _build_nettfront_order_import_csv(rows: list[NettfrontOrderRow]) -> bytes:
    """Build nettfront order import csv data."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\n")
    for row in rows:
        if _order_safe_number(row.order_qty) <= 0:
            continue
        part_number = _nettfront_order_display_part_number(row.part_number) or row.part_number.strip()
        if not part_number:
            continue
        writer.writerow([part_number, _nettfront_order_quantity_text(row.order_qty)])
    return buffer.getvalue().encode("utf-8-sig")


def _write_nettfront_order_bundle(job_dir: Path, metadata: dict) -> None:
    """Write nettfront order bundle data."""
    bundle_name = str(metadata.get("bundle_name", "nettfront-rendeles-output.zip")).strip() or "nettfront-rendeles-output.zip"
    bundle_files: list[str] = ["metadata.json", "suggestions.json", "rendelesi-javaslat.xlsx"]

    source_stock_file = str(metadata.get("source_stock_file", "")).strip()
    if source_stock_file:
        bundle_files.append(source_stock_file)

    source_parts_file = str(metadata.get("source_parts_file", "")).strip()
    if source_parts_file:
        bundle_files.append(source_parts_file)

    source_avg_file = str(metadata.get("source_average_file", "")).strip()
    if source_avg_file:
        bundle_files.append(source_avg_file)

    approved_file = str(metadata.get("approved_file", "")).strip()
    if approved_file:
        bundle_files.append(approved_file)

    import_file = str(metadata.get("import_file", "")).strip()
    if import_file:
        bundle_files.append(import_file)

    seen: set[str] = set()
    existing_files = []
    for file_name in bundle_files:
        if file_name in seen:
            continue
        seen.add(file_name)
        if (job_dir / file_name).exists():
            existing_files.append(file_name)

    (job_dir / bundle_name).write_bytes(create_bundle_archive(job_dir, existing_files))


def _write_nettfront_order_job(
    result,
    stock_name: str,
    stock_bytes: bytes,
    parts_name: str = "",
    parts_bytes: bytes | None = None,
    parts_count: int = 0,
) -> tuple[str, dict]:
    """Write nettfront order job data."""
    job_id = uuid.uuid4().hex[:12]
    job_dir = order_runtime_dir() / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    stock_suffix = Path(stock_name).suffix.lower() or ".xlsx"
    source_stock_file = f"source-stock{stock_suffix}"
    (job_dir / source_stock_file).write_bytes(stock_bytes)
    (job_dir / "rendelesi-javaslat.xlsx").write_bytes(result.suggestion_workbook)
    _write_nettfront_order_rows(job_dir, result.rows)

    metadata = {
        "job_id": job_id,
        "job_type": "order",
        "bundle_name": "nettfront-rendeles-output.zip",
        "source_stock_name": stock_name,
        "source_stock_file": source_stock_file,
        "suggestion_row_count": len(result.rows),
        "merged_variant_count": result.merged_variant_count,
        "filtered_stock_count": result.filtered_stock_count,
        "added_super_matt_count": result.added_super_matt_count,
        "total_m2": result.total_m2,
        "avg_row_count": result.avg_row_count,
        "approved_row_count": 0,
        "approved_total_m2": 0.0,
        "approved_file": "",
        "approved_generated_at": "",
    }

    if parts_name and parts_bytes is not None:
        parts_suffix = Path(parts_name).suffix.lower() or ".xlsx"
        parts_file = f"source-parts{parts_suffix}"
        (job_dir / parts_file).write_bytes(parts_bytes)
        metadata["source_parts_name"] = parts_name
        metadata["source_parts_file"] = parts_file
        metadata["source_parts_count"] = max(0, int(parts_count))

    (job_dir / "metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_nettfront_order_bundle(job_dir, metadata)
    return job_id, metadata


def _persist_nettfront_order_approval(job_dir: Path, metadata: dict, rows: list[NettfrontOrderRow]) -> dict:
    """Handle persist nettfront order approval logic for the NettFront workflows."""
    suggestion_workbook = rows_to_suggestion_workbook(rows)
    approved_title = f"Divian-Mega Kft. Rendelés {datetime.now().strftime('%Y.%m.%d.')}"
    approved_workbook = rows_to_approved_workbook(rows, approved_title)
    import_csv = _build_nettfront_order_import_csv(rows)

    (job_dir / "rendelesi-javaslat.xlsx").write_bytes(suggestion_workbook)
    (job_dir / "rendeles-jovahagyott.xlsx").write_bytes(approved_workbook)
    (job_dir / "rendeles_sima.csv").write_bytes(import_csv)
    _write_nettfront_order_rows(job_dir, rows)

    updated_metadata = {
        **metadata,
        "suggestion_row_count": len(rows),
        "total_m2": calc_total_m2_from_rows(rows),
        "approved_row_count": _count_positive_order_rows(rows),
        "approved_total_m2": calc_total_m2_from_rows(rows),
        "approved_file": "rendeles-jovahagyott.xlsx",
        "import_file": "rendeles_sima.csv",
        "approved_generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    (job_dir / "metadata.json").write_text(json.dumps(updated_metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_nettfront_order_bundle(job_dir, updated_metadata)
    return updated_metadata
